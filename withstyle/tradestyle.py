from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# from openpyxl.utils import coordinate_from_string, column_index_from_string

from openpyxl.styles import (PatternFill,  Border, Side, 
                             Alignment, Font, NamedStyle)
# Currently unused styles GradientFill, Protection, colors, Color,  
from withstyle.thetradeobject import SumReqFields

#global variable
srf = SumReqFields()

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
#         ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def c(icell, end=None, anchor= None) :
    '''
    Translates numerical coordinates as tuplits to excels  string addresses e.g., (1,1) returns 'A1' 
    :params:icell: The cell to translate.
    :parmas:end: Creates an excel range with icell at the top left and end at the bottom right.
    :params:anchor: A translation amount. The cell or range  will be translated by the amount of the cell -1
                    to account for all cells beginning at (1,1) not(0,0)
    '''
    if anchor :
        anchor = (anchor[0] - 1, anchor[1] - 1)
        icell = (icell[0] + anchor[0], icell[1] + anchor[1])
        if end :
            end = (end [0] + anchor[0], end[1] + anchor[1])

    # mini recursion to handle range
    if end is not None:
        return "{0}:{1}".format(c(icell),c(end))
    # For the single coordinate
    return "{0}{1}".format(get_column_letter(icell[0]),icell[1])
    

class TradeFormat (object):
    ''' Create the shape of the trade summary in an excel file and apply the styles defined here and named in for each cell in SumReqFields.tfcolumns'''
    # TODO openpyxl throws an exception if you name a named style twice.  Do something to deal with it. Either catch the exception or check if the style exists or 
    # make this a singleton (---no singleton, it doesn't address the issue)
    def __init__(self, wb, a=(1,1) ):
        self.tradeAnchor = a
        titleStyle = NamedStyle(name="titleStyle")
        titleStyle.font = Font(color="FFFFFF", size=16)
        titleStyle.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        titleStyle.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        titleStyle.border = Border(left=Side(style='double'),
                                 right=Side(style='double'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        titleStyle.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        wb.add_named_style(titleStyle) #This will be done automatically if you use it...
        
        
        titleStyleStratNotes = NamedStyle(name="titleStyleStratNotes")
        titleStyleStratNotes.font = Font(color="FFFFFF", size=8)
        titleStyleStratNotes.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        titleStyleStratNotes.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        titleStyleStratNotes.border = Border(left=Side(style='double'),
                                 right=Side(style='double'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        titleStyleStratNotes.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        wb.add_named_style(titleStyleStratNotes) #This will be done automatically if you use it...
        
        
       

        noteStyle = NamedStyle(name="noteStyle")
        noteStyle.font = Font(color="000000", size=10)
        noteStyle.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
        noteStyle.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        noteStyle.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        wb.add_named_style (noteStyle)
        
        
        normStyle = NamedStyle(name = "normStyle")
        normStyle.font = Font(color="FFFFFF", size=11)
        normStyle.alignment = Alignment(horizontal="left", vertical="bottom")
        normStyle.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normStyle.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        normStyle.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        wb.add_named_style (normStyle)
       
       
        normStyleRR = NamedStyle(name = "normStyleRR")
        normStyleRR.font = Font(color="FFFFFF", size=11)
        normStyleRR.alignment = Alignment(horizontal="left", vertical="bottom")
        normStyleRR.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normStyleRR.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        normStyleRR.number_format = '# ??/??'
        wb.add_named_style (normStyleRR)
        
        # TODO Have not created a seperate titleStyleLink style, using titleStyle
        self.styles = dict(
            titleStyle = titleStyle,
            titleStyleStratNotes = titleStyleStratNotes,
            titleStyleLink = titleStyle,
            normStyle = normStyle,
            normStyleRR = normStyleRR,
            noteStyle  = noteStyle
            
        )

    def mergeStuff(self, ws, icell, end, anchor=None) :
        if anchor :
            anchor = (anchor[0] - 1, anchor[1] - 1)
            icell = (icell[0] + anchor[0], icell[1] + anchor[1])
            end = (end [0] + anchor[0], end[1] + anchor[1])

        ws.merge_cells(start_row=icell[1], 
                       start_column=icell[0], 
                       end_row=end[1], 
                       end_column=end[0])

        return (icell, end)
    

    def popTheTrade(self, ws, tf, row):
        anchor = "A:" + row
        srf.tfcolumns
        
        
    def formatTrade(self, ws, tstyle=None, anchor = None) :
        if anchor:
            self.tradeAnchor = anchor
        anc = self.tradeAnchor
        
        for val in  srf.tfcolumns.values() :
            if isinstance(val[0], list) :
                self.mergeStuff(ws, val[0][0], val[0][1], anchor=anc)
                rng = c(val[0][0], val[0][1], anchor=anc)
                
                style_range(ws, rng, self.styles[val[1]].border)
                ws[c(val[0][0], anchor=anc)].style = self.styles[val[1]]

            else :
                ws[c(val[0], anchor=anc)].style = self.styles[val[1]]
        
    

        
        # Leave this in till we get till the good stuff gets as a visual check we be on track

        
        
        ws[c((7,5), anchor=anc)]  = "(Explanatory note for strategy)"
        ws[c((1,11), anchor=anc)] = "(Technical description of the trade.)"
        ws[c((1,17), anchor=anc)] = "(Evaluation of the trade)"
    



  
# 
#     def c(self, icell,  add= None) :
#         if add :
#             icell = (icell[0] + add[0], icell[1] + add[1])
#         return "{0}{1}".format(get_column_letter(icell[0]),icell[1])
    
   
    
   

# 
# wb=Workbook()
# 
# ws = wb.active
# tf = TradeFormat(wb, a = ((1,45)))
# tf.formatTrade(ws)
# tf.formatTrade(ws, anchor=(7,96))
# 
# 
# wb.save('out/justlookin.xlsx')
# print("done!")












