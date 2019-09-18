# Structjour -- a daily trade review helper
# Copyright (C) 2019 Zero Substance Trading
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.

'''
Test the methods in the module layoutsheet

@created_on Feb 8, 2019

@author: Mike Petersen
'''

import datetime as dt
import os
from random import randint
from unittest import TestCase
import unittest
from unittest.mock import patch
from collections import deque

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

from PyQt5.QtCore import QSettings

from structjour.layoutsheet import LayoutSheet
from structjour.tradestyle import TradeFormat, c as tcell
########: disable = C0103, W0613, W0603, W0212, R0914
# pylint: disable = C0103, W0613, W0603, W0212, R0914


D = deque()


# Runs w/o error using discovery IFF its the first test to run. With mock, QTest, global data and 
# precarious data matching based on knowing too much, I am not going to take the time to fix it.
# Its deprecated code- will soon remove all evidence of the console version and the cosole
# interview. (It runs without error when run seperately)
class TestLayoutSheet(TestCase):
    '''
    Run all of structjour with a collection of input files and test the outcome
    '''

    def __init__(self, *args, **kwargs):
        super(TestLayoutSheet, self).__init__(*args, **kwargs)

        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

        settings = QSettings('zero_substance', 'structjour')
        settings.setValue('runtype', 'CONSOLE')

        self.dadata = deque(
            [[-4000], [3923, -750], [], [600, 241, 50], [-169],
             [], [0, -600], [], [0, 750, 50], [-600], [0, -200]])

        # Input test files can be added here. And place the test data in testdata.xlsx. Should add
        # files with potential difficulties
        self.infiles = ['trades.1116_messedUpTradeSummary10.csv', 'trades.8.WithHolds.csv',
                        'trades.8.csv', 'trades.907.WithChangingHolds.csv',
                        'trades_190117_HoldError.csv', 'trades.8.ExcelEdited.csv',
                        'trades.910.tickets.csv', 'trades_tuesday_1121_DivBy0_bug.csv',
                        'trades.8.WithBothHolds.csv', 'trades1105HoldShortEnd.csv',
                        'trades190221.BHoldPreExit.csv']

        # self.tests = self.getTestData(r'C:\python\E\structjour\src\data')

    def setUp(self):

        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

    def getTestData(self, indir):
        '''
        Open the csv file testdata and oraganze the data into a usable data structure. The file is
        necessarily populated by 'hand.' To add a file to test, copy it to the data dir and enter
        the information to test.
        :return data: List  containing the data to check against the output files.
        '''

        df = pd.read_excel(os.path.join(indir, 'testdata.xlsx'))

        l = len(df)
        data = list()

        data = list()
        for i, row in df.iterrows():
            entry = list()
            if not pd.isnull(df.at[i, 'Order']):
                entry.extend((row['Order'], row['NumTrades'], row['Name']))
                j = i
                trades = list()
                begginning = True
                while j < l and not pd.isnull(df.at[j, 'Ticker']):
                    # Check these specific trades
                    if not begginning:
                        if isinstance(df.at[j, 'Name'], str):
                            break

                    trades.append([df.at[j, 'Ticker'], df.at[j, 'Account'],
                                   df.at[j, 'Side'], df.at[j, 'Held'], df.at[j, 'Pos'], ])
                    begginning = False
                    j = j+1
                entry.append(trades)
                data.append(entry)
        return data

    def test_createWorkbook(self):
        '''
        Test the method structjour.layoutsheet.LayoutSheet.createWorkbook
        '''

        df = pd.DataFrame(np.random.randint(0, 100, size=(100, 7)), columns=list('ABCDEFG'))
        # df
        margin = 25
        spacing = 3
        inputlen = len(df)
        ls = LayoutSheet(margin, inputlen, spacing=spacing)

        wb, ws, df = ls.createWorkbook(df)


        for row, (i, dfrow) in zip(ws, df.iterrows()):
            # We inserted the column headers in this row (ws starts with 1, not 0)
            if i + 1 == ls.topMargin:
                for ms, x in zip(row, df.columns):
                    self.assertEqual(x, ms.value)
            # everything else is verbatim
            else:
                for ms, x in zip(row, dfrow):
                    self.assertEqual(x, ms.value)


        wb.save("out/SCHNOrK.xlsx")

    def test_styleTop(self):
        '''
        Test the method layoutsheet.LayoutSheet.styleTop. This  will probably produce warnings from
        openpyxl as there is empty data when it makes the headers. No worries.
        Note that we are using a protected member of Worksheet ws._tables, so if it this fails, look
        at that. openpyxl does not provide a public attribute for tables.
        Note that knowing the quoteRange and noteRange is bad design. Eventually these two bits of
        design data should be abstracted to somewhere accessible by the user. (and testing too)
        '''
        quoteRange = [(1, 1), (13, 5)]
        noteRange = [(1, 6), (13, 24)]
        quoteStyle = 'normStyle'
        noteStyle = 'explain'
        margin = 25
        inputlen = 50   #len(df)

        wb = Workbook()
        ws = wb.active

        # Make sure the out dir exists
        if not os.path.exists("out/"):
            os.mkdir("out/")

        # Make sure the file we are about to create does not exist
        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        # Create table header and data in the ws
        headers = ['Its', 'the', 'end', 'of', 'the', 'world', 'as', 'we',
                   'know', 'it.', 'Bout', 'Fn', 'Time!']
        for i in range(1, 14):
            ws[tcell((i, 25))] = headers[i-1]

        ls = LayoutSheet(margin, inputlen)
        for x in range(ls.topMargin+1, ls.inputlen + ls.topMargin+1):
            for xx in range(1, 14):
                ws[tcell((xx, x))] = randint(-1000, 10000)


        ls.styleTop(ws, 13, TradeFormat(wb))

        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        listOfMerged = list()
        listOfMerged.append(tcell((quoteRange[0])) + ':' +  tcell((quoteRange[1])))
        listOfMerged.append(tcell((noteRange[0])) + ':' +  tcell((noteRange[1])))
        for xx in ws2.merged_cells.ranges:
            # print (str(xx) in listOfMerged)
            self.assertTrue(str(xx) in listOfMerged)
        self.assertEqual(ws[tcell(quoteRange[0])].style, quoteStyle)
        self.assertEqual(ws[tcell(noteRange[0])].style, noteStyle)

        self.assertEqual(len(ws._tables), 1)


        begin = tcell((1, ls.topMargin))

        end = tcell((13, ls.topMargin + ls.inputlen))
        tabRange = f'{begin}:{end}'
        self.assertEqual(tabRange, ws._tables[0].ref)

        os.remove(dispath)

    def test_styleTopwithnothin(self):
        '''
        Test the method layoutsheet.LayoutSheet.styleTop. Test that it still works without
        table data. We still know too much about the method,. Note that we are using a protected
        member of Worksheet ws._tables
        '''
        quoteRange = [(1, 1), (13, 5)]
        noteRange = [(1, 6), (13, 24)]
        quoteStyle = 'normStyle'
        noteStyle = 'explain'
        margin = 25
        inputlen = 50   #len(df)

        wb = Workbook()
        ws = wb.active

        # Make sure the out dir exists
        if not os.path.exists("out/"):
            os.mkdir("out/")

        # Make sure the file we are about to create does not exist
        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        ls = LayoutSheet(margin, inputlen)
        ls.styleTop(ws, 13, TradeFormat(wb))

        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        listOfMerged = list()
        listOfMerged.append(tcell((quoteRange[0])) + ':' +  tcell((quoteRange[1])))
        listOfMerged.append(tcell((noteRange[0])) + ':' +  tcell((noteRange[1])))
        for xx in ws2.merged_cells.ranges:
            # print (str(xx) in listOfMerged)
            self.assertIn(str(xx), listOfMerged)
        self.assertEqual(ws[tcell(quoteRange[0])].style, quoteStyle)
        self.assertEqual(ws[tcell(noteRange[0])].style, noteStyle)

        self.assertEqual(len(ws._tables), 1)


        begin = tcell((1, ls.topMargin))

        end = tcell((13, ls.topMargin + ls.inputlen))
        tabRange = f'{begin}:{end}'
        self.assertEqual(tabRange, ws._tables[0].ref)

        os.remove(dispath)




def notmain():
    '''Run some local code'''
        # pylint: disable = E1120
    ttt = TestLayoutSheet()
    ttt.test_createWorkbook()
    ttt.test_styleTop()
    ttt.test_styleTopwithnothin()

    # ttt.test_populateMistakeForm()
    # ttt.test_populateDailySummaryForm()
    # ttt.test_runSummaries()
    # ttt.test_createImageLocation()

def main():
    '''Run unittests cl style'''
    #import sys;sys.argv = ['', 'Test.testName']
    unittest.main()




if __name__ == '__main__':
    # notmain()
    main()
