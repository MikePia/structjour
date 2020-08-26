# Structjour -- a daily trade review helper
# Copyright (C) 2019 Zero Substance Trading

# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.

'''
Test the methods in the module exporttoexcel

@created_on Feb 8, 2019

@author: Mike Petersen
'''
import os
import sys

from unittest import TestCase
import unittest

from openpyxl import Workbook, load_workbook

from PyQt5.QtCore import QSettings

# from structjour.rtg import getRandGenTradeStuff


from structjour.dailysumforms import MistakeSummary
from structjour.definetrades import DefineTrades
from structjour.journalfiles import JournalFiles

from structjour.models.meta import ModelBase

from structjour.utilities.rtg import RTG
from structjour.statements.ibstatement import IbStatement
from structjour.statements.ibstatementdb import StatementDB
from structjour.statements.statement import getStatementType
from structjour.statements.dasstatement import DasStatement
from structjour.statements.dailynotescrud import DailyNotesCrud

from structjour.stock.utilities import clearTables
from structjour.tradestyle import c as tcell
from structjour.utilities.backup import Backup
from structjour.view.dailycontrol import DailyControl
from structjour.view.exportexcel import ExportToExcel
from structjour.view.layoutforms import LayoutForms
from structjour.view.sumcontrol import SumControl

# from PyQt5.QtTest import QTest
from PyQt5 import QtCore
QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_ShareOpenGLContexts)
from PyQt5 import QtWebEngineWidgets    # noqa:  F401
from PyQt5.QtWidgets import QApplication

# pylint: disable = C0103

app = QApplication(sys.argv)


class Test_ExportToExcel(TestCase):
    '''
    Test the ExportToExcel Object. Generally use this class for less elaborate setups and
    Test_ExportToExcel_MistakeData for more elaborate setups (Hyperinks and stuff requiring
    coordinated data)
    '''

    infiles = []
    dates = []
    lfs = []
    infiles = []
    dates = []
    jfs = []
    dframes = []
    ldfs = []

    rtg = None
    db = ''
    outdir = ''
    sc = None

    @classmethod
    def setUpClass(cls):
        bu = Backup()
        bu.backup()
        if ModelBase.session:
            ModelBase.session.rollback()
        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

        outdir = 'test/out'
        cls.outdir = os.path.realpath(outdir)
        cls.db = 'data/testdb.sqlite'
        cls.db = os.path.realpath(cls.db)
        if os.path.exists(cls.db):
            clearTables(cls.db)

        cls.rtg = RTG(db=cls.db)
        # cls.dates = ['20200203 09:30', '20200204 07:30', '20200205 09:35', '20200206 11:40', '20200207 10:39']
        cls.dates = ['20200207 10:39']
        cls.infiles = cls.rtg.saveSomeTestFiles(cls.dates, cls.outdir, strict=True, overwrite=True)

        settings = QSettings('zero_substance', 'structjour')
        for i, name in enumerate(cls.infiles):
            name = os.path.join(cls.outdir, name)
            x, cls.inputType = getStatementType(name)
            # print(cls.inputType)
            if cls.inputType == 'DAS':
                ds = DasStatement(name, settings, cls.dates[i])
                ds.getTrades(testFileLoc=cls.outdir, testdb=cls.db)
            elif cls.inputType == "IB_CSV":
                ibs = IbStatement(db=cls.db)
                ibs.openIBStatement(name)
            else:
                continue
            #     self.assertTrue(4 == 5, "Unsupported file type in test_TheTradeObject")

            statement = StatementDB(db=cls.db)
            df = statement.getStatement(cls.dates[i])
            # self.assertFalse(df.empty, f"Found no trades in db on {daDate}")
            dtrade = DefineTrades(cls.inputType)
            dframe, ldf = dtrade.processDBTrades(df)
            # tto = TheTradeObject(ldf[0], False, SumReqFields())
            jf = JournalFiles(indir=cls.outdir, outdir=outdir, theDate=cls.dates[i], infile=name)
            cls.sc = SumControl()
            lf = LayoutForms(cls.sc, jf, dframe)
            lf.runTtoSummaries(ldf)
            cls.jfs.append(jf)
            cls.dframes.append(dframe)
            # cls.ttos.append(tto)
            cls.ldfs.append(ldf)
            cls.lfs.append(lf)
        # rw = runController(w)

    @classmethod
    def tearDownClass(cls):
        bu = Backup()
        bu.restore()

    def test_exportExcel(self):
        '''Using a random trade generator, test exportExcel creates a file'''

        for jf, lf, df in zip(self.jfs, self.lfs, self.dframes):

            t = ExportToExcel(lf.ts, jf, df)
            t.exportExcel()

            self.assertTrue(os.path.exists(jf.outpathfile))

    def test_populateXLDailyNote(self):
        '''
        Test the method populateXLDailyNote. Specifically test after setting a note, calling the
        method the workbook contains the note as expected
        '''
        for jf, lf, df in zip(self.jfs, self.lfs, self.dframes):
            dnc = DailyNotesCrud(jf.theDate)
            note = 'Ipsum solarium magus coffeum brewum'
            dnc.setNote(note)

            t = ExportToExcel(lf.ts, jf, df)
            wb = Workbook()
            ws = wb.active
            t.populateXLDailyNote(ws)
            t.saveXL(wb, jf)
            wb2 = load_workbook(jf.outpathfile)

            cell = (1, 6)
            cell = tcell(cell)
            val = wb2.active[cell].value
            self.assertEqual(note, val)

    def test_populateXLDailySummaryForm(self):
        '''
        Test the method populateXLDailySummaryForm. Specifically test that given a specific anchor,
        12 specific spots are written to by default.
        '''
        for jf, lf, df, ldf in zip(self.jfs, self.lfs, self.dframes, self.ldfs):
            NUMROWS = 6
            mstkAnchor = (3, 3)     # arbitrary
            mistake = MistakeSummary(numTrades=len(ldf), anchor=mstkAnchor)
            t = ExportToExcel(lf.ts, jf, df)
            wb = Workbook()

            t.populateXLDailySummaryForm(mistake, wb.active, mstkAnchor)
            # t.saveXL(wb, self.jf)

            anchor = (mstkAnchor[0], mstkAnchor[1] + len(ldf) + 5)
            cell = tcell(mstkAnchor, anchor=anchor)
            for i in range(0, NUMROWS):
                cell = tcell((mstkAnchor[0], mstkAnchor[1] + i), anchor=anchor)
                cell2 = tcell((mstkAnchor[0] + 1, mstkAnchor[1] + i), anchor=anchor)
                self.assertIsInstance(wb.active[cell].value, str)
                self.assertIsInstance(wb.active[cell2].value, str)


def main():
    unittest.main()


def notmain():

    Test_ExportToExcel.setUpClass()
    t = Test_ExportToExcel()

    # t.test_exportExcel()
    # t.test_populateXLDailyNote()
    t.test_populateXLDailySummaryForm()


if __name__ == '__main__':
    # notmain()
    main()
