import datetime, os


from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from journalfiles import JournalFiles
from structjour.pandasutil import DataFrameUtil, InputDataFrame, ToCSV_Ticket as Ticket
from structjour.tradeutil import ReqCol, FinReqCol, XLImage, TradeUtil
from withstyle.thetradeobject import SumReqFields, TheTradeObject
from withstyle.tradestyle import TradeFormat
from withstyle.tradestyle import c as tcell

# jf = JournalFiles(indir= "C:\trader\journal\_08_August\Week_5\_0831_Friday",mydevel=True)

# jf= JournalFiles(indir =r'C:\trader\journal\_09_September\Week_1\_0904_Tuesday', infile='trades1.csv', mydevel=True)
# jf = JournalFiles(theDate=datetime.date(2018, 9,6), outdir = 'out', mydevel=True)
# jf = JournalFiles(indir='data', infile='TradesWithHolds.csv', outdir = "out", mydevel=True)
# jf = JournalFiles(theDate = datetime.date(2018, 10, 1), outdir = 'out/', mydevel = True)
#jf=JournalFiles(outdir='out/', mydevel=True)
jf = JournalFiles(mydevel = True, outdir='out/')
jf._printValues()
        
tkt = Ticket(jf)

tu = TradeUtil()
trades, jf =tkt.newDFSingleTxPerTicket()
# trades = pd.read_csv(jf.inpathfile)

idf = InputDataFrame()
reqCol = ReqCol()
finalReqCol = FinReqCol()

#Process the input file DataFrame
DataFrameUtil.checkRequiredInputFields(trades, reqCol.columns)
trades = idf.zeroPadTimeStr(trades)
trades = trades.sort_values([reqCol.acct, reqCol.ticker, reqCol.time])
trades = idf.mkShortsNegative(trades)
swingTrade = idf.getOvernightTrades(trades)
swingTrade = idf.figureOvernightTransactions(trades)
trades = idf.insertOvernightRow(trades,swingTrade)

#Process the output file DataFrame
trades = tu.addFinReqCol(trades)
newTrades = trades[finalReqCol.columns]
newTrades.copy()
nt = newTrades.sort_values([finalReqCol.ticker,finalReqCol.acct,  finalReqCol.time])
nt = tu.writeShareBalance(nt)
nt = tu.addStartTime(nt)
nt = nt.sort_values([finalReqCol.start, finalReqCol.acct, finalReqCol.time])
nt = tu.addTradeIndex(nt)
nt = tu.addTradePL(nt)
nt = tu.addTradeDuration(nt)
nt = tu.addTradeName(nt)
nt=DataFrameUtil.addRows(nt,1)
nt = tu.addSummaryPL(nt)
ldf=tu.getTradeList(nt)
inputlen = len(nt)              # Get the length of the input file in order to style it in the Workbook
dframe = DataFrameUtil.addRows(nt, 2)



#Process the openpyxl excel object using the output file DataFrame. Insert images and Trade Summaries.
topMargin = 10
newdf = DataFrameUtil.createDf(dframe,  topMargin)
insertsize = 25
dframe = newdf.append(dframe, ignore_index = True)

#Add rows and append each trade, leaving space for an image. Create a list of names and row numbers 
# to place images within the excel file (imageLocation data structure).
imageLocation = list()
for tdf in ldf :
    imageName='{0}_{1}_{2}_{3}.jpeg'.format (tdf[finalReqCol.tix].unique()[-1].replace(' ',''), 
           tdf[finalReqCol.name].unique()[-1].replace(' ','-'),
           tdf[finalReqCol.start].unique()[-1],
           tdf[finalReqCol.dur].unique()[-1])

    # TODO handle empty string in the tdf
    imageLocation.append([len(tdf) + len(dframe) + 2, 
                          tdf.Tindex.unique()[0].replace(' ', '') + '.jpeg',
                          imageName,
                          tdf.Start.unique()[-1],
                        tdf.Duration.unique()[-1]])
    print(len(tdf) + len(dframe) + 2)

    dframe = dframe.append(tdf, ignore_index = True)
    dframe = DataFrameUtil.addRows(dframe, insertsize)
    print(len(dframe))

nt = dframe

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(nt, index=False, header=False):
    ws.append(r)

for name, cell  in zip(nt.columns, ws[topMargin]) :
    cell.value = name

#Style the table, and the top paragraph.  Add and style the inspire quote
tblRng= "{0}:{1}".format(tcell((1,topMargin)), tcell((len(nt.columns),topMargin+inputlen)))
tab = Table(displayName="Table1", ref=tblRng)
style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style

ws.add_table(tab)




XL = XLImage()

srf = SumReqFields()
tradeSummaries = list()
tf = TradeFormat(wb)
assert (len(ldf) == len(imageLocation))
     
response = input("Would you like to enter strategy names, targets and stops?")
interview = True if response.lower().startswith('y') else False

for loc, tdf in zip(imageLocation, ldf) :
#     print('Copy an image into the clipboard for {0} beginning {1}, and lasting {2}'.format(loc[1], loc[2], loc[3]))
    img = XL.getAndResizeImage(loc[2], jf.outdir)
    cellname = 'J' + str(loc[0])
    ws.add_image(img, cellname)
    
    #Put together the trade summary info for each trade and interview the trader
    tto=TheTradeObject(tdf, interview)
    tto.runSummary()
    tradeSummaries.append(tto)
    
    #Place the format shapes/styles in the worksheet
    tf.formatTrade(ws, anchor=(1, loc[0]))
    
    #populate the trade information
    for key in srf.tfcolumns.keys()  :
        cell = srf.tfcolumns[key][0]
        if isinstance(cell, list) :
            cell = cell[0]
        tradeval = tto.TheTrade[key].unique()[0]
    #     print ("{0:10} \t{3} \t{1:}\t{2} ".format(key, cell, tradeval, tcell(cell, anchor=(1, loc[0]))))


        # Put some formulas in each trade Summary
        if key in srf.tfformulas :
            
            anchor=(1,loc[0])
            formula=srf.tfformulas[key][0]
            args=[]
            for c in srf.tfformulas[key][1:] :
                args.append(tcell(c, anchor=anchor))
            tradeval = formula.format(*args)
            
        if not tradeval :
            continue
        ws[tcell(cell, anchor=(1, loc[0]))] = tradeval

    
print("Done with interview")



#Write the file
jf.mkOutdir() 
saveName=jf.outpathfile
count=1
while True :
    try :
        wb.save(saveName)
    except PermissionError as ex :
        print(ex)
        print("Failed to create file {0}.{1}".format(saveName, ex))
        print("Images from the clipboard were saved  in {0}".format(jf.outdir))
        (nm, ext) = os.path.splitext(jf.outpathfile)
        saveName = "{0}({1}){2}".format(nm,count,ext)
        print("Will try to save as {0}".format(saveName))
        count=count+1
        if count==6:
            print("Giving up. PermissionError")
            raise (PermissionError("Failed to create file {0}".format(saveName)))
        continue
    except Exception as ex:
        print (ex)
    break
print("Done!")
