import datetime
import os
from openpyxl import load_workbook


from journal.dfutil import DataFrameUtil
from journal.thetradeobject import SumReqFields
from journal.tradestyle import c as tcell

import pandas as pd
# pylint: disable=C0103


class DisReqCol(object):
    def __init__(self, thedate=datetime.date.today()):
        self.thedate = thedate
        rcvals = ["Date", "Time", "Long/Short", "Ticker", "Entry Price", "Account Balance", "Position",
                  "Stop Loss", "Target", "Average Exit", "P / L", "Strategy", "Trade Notes"]

        rckeys = ["date", "time", "side", "symb", "entry1", "acctbal", "shares",
                  "stoploss", "targ", "avgexit", "pl", "strat", "notes"]

        rc = dict(zip(rckeys, rcvals))

        self._date = rc['date']
        self._time = rc['time']
        self.side = rc['side']
        self.symb = rc['symb']
        self.entry1 = rc['entry1']
        self.acctbal = rc['acctbal']
        self.shares = rc['shares']
        self.stoploss = rc['stoploss']
        self.target = rc['targ']
        self.avgexit = rc['avgexit']
        self.pl = rc['pl']
        self.strat = rc['strat']
        self.notes = rc['notes']

        # We really don't need to manage the styles but just let these
        # sit here for uniformity-- or finding we actually do need them.
        tfcolumns = {
            "date": [(1, 1), 'dateStyle'],
            "time": [(2, 1), 'timeStyle'],
            "side": [(3, 1), 'normal'],
            "symb": [(4, 1), 'normal'],
            "entry1": [(5, 1), 'currencyStyle'],
            "acctbal": [(6, 1), 'currencyStyle'],
            "shares": [(7, 1), 'numberStyle'],
            "stoploss": [(8, 1), 'numberStyle'],
            "targ": [(9, 1), 'currencyStyle'],
            "avgexit": [(10, 1), 'currencyStyle'],
            "pl": [(11, 1), 'currencyStyle'],
            "strat": [(12, 1), 'normal'],
            "notes": [(16, 1), 'normal']}

        self.rc = rc
        self.tfcolumns = tfcolumns


def loadTradeSummaries(loc, trades):
    '''
    Load up each trade summary in the excel doc into a 1 row DataFrame, return a list
    of these DataFrames
    :params:loc: A list of the rows within the excel doc on which to find the trade summaries
    :return: A list of 1-row DataFrames Each trade is on one row from each of the trade summay forms
    '''

    ldf = list()
    srf = SumReqFields()
    reqCol = srf.rc
    newdf = pd.DataFrame(columns=reqCol.values())
    colFormat = srf.tfcolumns

    for rowNum in loc:
        newdf = DataFrameUtil.createDf(newdf, 1)
        for key in reqCol.keys():
            cell = colFormat[reqCol[key]][0]
            if isinstance(cell, list):
                cell = cell[0]
            cell = tcell(cell, anchor=(1, rowNum))
            newdf.iloc[-1][reqCol[key]] = trades[cell].value
        ldf.append(newdf)
    return ldf


def getTradeSummaryFormLocations(ws):
    '''
    Find the Trade Summary Forms within the excel doc and save their location in tradeLoc 
    trades.max_row
    :params:ws: The openpyxl Worksheet object to search
    '''
    loc = list()

    #Search for trade names like 'AAPL Short' in column A. Save the cell for each to loc
    for row in ws.iter_rows():
        val = row[0].value
        if val and len(val) > 3 and len(val) < 15:
            if 'short' in val.lower() or 'long' in val.lower():
                r = row[0].row
                loc.append(r)
    return loc


def getAvgExit(trade):
    '''We want to display jut one exit in the Trade log. It might as well be the average
    exit when we scale out. If there are more than 8 combined entries and exits--could be
    a problem
    TODO: track down the note for more entrances/exits than 8 -- It may be complicated-- 
    entered a bug--classic procrastination
    '''
    product = 0
    exits = list()
    positions = list()
    for i in range(1, 9):
        ex = "Exit{}".format(i)
        shares = "EShare{}".format(i)
        if trade[ex].unique()[0]:
            x = trade[ex].unique()[0]
            p = trade[shares].unique()[0]
            exits.append(x)
            positions.append(p)
            try:
                assert(positions[-1])
            except:
                print(
                    "Found empty shares in an exit-- this may be an overnight trade or something: Fix this bug developer!!")
            product = product + (x * p)

    return 0 if sum(positions) == 0 else product/sum(positions)


def gotAnyExits(trade):
    '''
    Return True if this trade processed at least one exit. Return False if the exits are all None
    :params:trade: DataFrame representing a single trade and with columns Exit1, Exit2 ... Exit8
    '''
    for i in range(1, 9):
        ex = "Exit{}".format(i)
        if trade[ex].unique()[0]:
            return True
    return False


def getWeekCount(theDate):
    '''
    Returns the number of weeks that have happend in the month counting only weekdays. If the first day
    of the month is Friday, the following Monday begins week 2. If the first day of the week is Saturday
    Week 2 will begin on Monday the 10th.
    :params:theDate: the date to test. A timedate.date object
    '''
    startMonth = datetime.date(theDate.year, theDate.month, 1)
    delt = datetime.timedelta(1)
    weekCount = 1
    startCounting = False
    while startMonth <= theDate:

        if(startMonth.weekday() == 4):
            startCounting = True
        if startCounting == True and startMonth.weekday() == 0:
            weekCount = weekCount + 1
#         print(startMonth, weekCount, startMonth.weekday(), startMonth.strftime("%A"))
        startMonth = startMonth+delt
    return weekCount


def getDevelDailyJournalList(prefix, begin):
    '''
    Gets a list of Daily trade files as created by Structjour in the MyDevel directories.
    :params:prefix: The directory that holds all the journal files.
    :params:date: A datetime.date object identifying the earliest file. All files up to today 
                    are in the list with the date of the file 
    :return: A list of filenames with dates[[filename, date], [filename2, date2]]...
    '''
    thelist = list()
    # prefix=prefix
    weekCount = getWeekCount(begin)
    now = datetime.date.today()
    theDate = begin
    delt = datetime.timedelta(1)
    oldmonth = theDate.month

    while now >= theDate:
        newmonth = theDate.month
        if theDate.weekday() == 6:
            weekCount = weekCount + 1
        if newmonth != oldmonth:
            weekCount = 1
        oldmonth = newmonth

        filename = "{}{}/Week_{}/{}".format(prefix,
                                            theDate.strftime("_%m_%B"),
                                            weekCount,
                                            theDate.strftime(
                                                "_%m%d_%A/out/Trades_%A_%m%d.xlsx")
                                            )
        filename = os.path.normpath(filename)
        if theDate.weekday() < 5:
            if not os.path.exists(filename):
                print("no file for ", filename)
            else:
                thelist.append([filename, theDate])
        theDate = theDate+delt

    return thelist


def registerTrades(tsList, wb):
    for fname, theDate in tsList:
        print(fname, theDate)
        wb2 = load_workbook(fname)
        trades = wb2["Sheet"]

        tradeLocation = getTradeSummaryFormLocations(trades)

        ldf = loadTradeSummaries(tradeLocation, trades)
        drc = DisReqCol(theDate)

        tlog = wb["Trade Log"]

        startSearch = False
        ix = -2
        cols = drc.tfcolumns
        # Here is a list of the keys to use cols.keys() of the trade log DataFrame
        #['date', 'time', 'side', 'symb', 'entry1', 'acctbal', 'shares',
        #'stoploss', 'targ', 'avgexit', 'pl', 'notes'])
        # Not bothering with the abstraction (drc.name) because this is entirely ours.

        srf = SumReqFields()

        for row in tlog.iter_rows():
            anchor = (1, row[0].row)

            if startSearch == True:
                if not row[0].value:
                    startSearch = False
                    ix = 0
            if row[0].value == 'Date':
                startSearch = True

            if ix >= 0:
                tdf = ldf[ix]
                if not gotAnyExits(tdf):
                    continue
                # print(cols.keys())

                #date
                cell = tcell(cols['date'][0], anchor=anchor)
                tlog[cell] = theDate

                #time
                cell = tcell(cols['time'][0], anchor=anchor)
                tim = tdf[srf.start].unique()[0]
                tlog[cell] = tim

                #side
                name = tdf[srf.name].unique()[0]
                if name:
                    cell = tcell(cols['side'][0], anchor=anchor)
                    tlog[cell] = name.split()[1]

                #symb
                cell = tcell(cols['symb'][0], anchor=anchor)
                tlog[cell] = name.split()[0]

                #entry1
                cell = tcell(cols['entry1'][0], anchor=anchor)
                tlog[cell] = tdf[srf.entry1].unique()[0]

                #Account Balance (setting an excel formula)
                cell = tcell(cols['acctbal'][0], anchor=anchor)
                formula = "=$M$3+SUM($N$7:$N{})".format(row[0].row-1)
                tlog[cell] = formula

                # "shares"
                cell = tcell(cols['shares'][0], anchor=anchor)
                shares = tdf[srf.shares].unique()[0].split()[0]
                if len(shares) > 0:
                    try:
                        ishares = int(shares)
                    except:
                        ishares = 0
                tlog[cell] = ishares

                #stoploss
                cell = tcell(cols['stoploss'][0], anchor=anchor)
                sl = tdf[srf.stoploss].unique()[0]
                tlog[cell] = sl

                #target
                cell = tcell(cols['targ'][0], anchor=anchor)
                target = tdf[srf.targ].unique()[0]
                tlog[cell] = target

                #avgExit
                cell = tcell(cols['avgexit'][0], anchor=anchor)
                tlog[cell] = getAvgExit(tdf)

                # P/L
                cell = tcell(cols['pl'][0], anchor=anchor)
                pl = tdf[srf.pl].unique()[0]
                tlog[cell] = pl

                # Strategy
                cell = tcell(cols['strat'][0], anchor=anchor)
                strat = tdf[srf.strat].unique()[0]
                tlog[cell] = strat


                # notes (from the mistake note field)
                cell = tcell(cols['notes'][0], anchor=anchor)
                note = tdf[srf.mstknote].unique()[0]
                tlog[cell] = note

                ix = ix + 1
                if ix == len(ldf):
                    break


disPath = os.path.normpath("Disciplined.xlsx")
if os.path.exists(disPath):
    wb = load_workbook(disPath)
else:
    print("Does not exist", disPath)
    quit(0)


begin = datetime.date(2018, 10, 15)
prefix = "C:/trader/journal/"
flist = getDevelDailyJournalList(prefix, begin)
registerTrades(flist, wb)
wb.save(disPath)
print('done!')


# wb.save
# wb.save(disPath)
# disPath = os.path.normpath("Disciplined.xlsx")
# if os.path.exists(disPath) :
#     wb = load_workbook(disPath)
#     print("so far...")
# else:
#     print("Does not exist", disPath)
#     quit(0)

# begin=datetime.date(2018, 10, 15)
# prefix="C:/trader/journal/"
# flist = getDevelDailyJournalList(prefix, begin)
# for x in flist :
#     print(x)
# print('done!')
