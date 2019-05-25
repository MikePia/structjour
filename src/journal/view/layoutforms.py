'''
Populate the Qt forms. Generally layoutforms populates the trade summaries for QT and layoutsheets
populates the form entries for excel.

Created on April 14, 2019

@author: Mike Petersen
'''

import datetime as dt
import os
import pickle
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors

import numpy as np
import pandas as pd

from PyQt5.QtCore import QSettings

from journal.view.sumcontrol import qtime2pd

from journal.dfutil import DataFrameUtil
from journal.definetrades import FinReqCol, DefineTrades
from journal.thetradeobject import SumReqFields, TheTradeObject
from journal.tradestyle import c as tcell
from journal.layoutsheet import LayoutSheet
from journal.dailysumforms import MistakeSummary
from journal.view.dailycontrol import DailyControl


from journal.tradestyle import TradeFormat


# from journal.view.sumcontrol import SumControl
# pylint: disable=C0103, C1801


class LayoutForms:
    '''
    Run theTradeObject summaries to get the tto DataFrame and populate
    the trade form for each trade.
    '''

    def __init__(self, sc, jf, df):
        '''
        Initialize the obect and create a Widget dictionary using the same keys as the keys to the
        tto object to populate the data.
        '''
        self.jf = jf
        self.df = df
        if self.df is not None:
            self.pickleitnow()
        self.ts = dict()
        self.entries = dict()
        rc = SumReqFields()
        self.timeFormat = '%H:%M:%S'
        self.topMargin = 25             # For use in Excel export

        self.sc = sc
        self.tradeSummaries = None

        # Widget Dictionary. Keys are same keys for TheTradeObject.TheTrade object
        wd = dict()
        wd[rc.name] = sc.ui.title
        wd[rc.acct] = sc.ui.account
        wd[rc.strat] = sc.ui.strategy
        wd[rc.link1] = sc.ui.link
        wd[rc.pl] = sc.ui.pl
        wd[rc.start] = sc.ui.start
        wd[rc.dur] = sc.ui.dur
        wd[rc.shares] = sc.ui.pos
        wd[rc.mktval] = sc.ui.mkt
        wd[rc.targ] = sc.ui.targ
        wd[rc.targdiff] = sc.ui.targDiff
        wd[rc.stoploss] = sc.ui.stop
        wd[rc.sldiff] = sc.ui.stopDiff
        wd[rc.rr] = sc.ui.rr
        wd[rc.maxloss] = sc.ui.maxLoss
        wd[rc.mstkval] = sc.ui.lost
        wd[rc.mstknote] = sc.ui.sumNote
        wd[rc.entry1] = sc.ui.entry1
        wd[rc.entry2] = sc.ui.entry2
        wd[rc.entry3] = sc.ui.entry3
        wd[rc.entry4] = sc.ui.entry4
        wd[rc.entry5] = sc.ui.entry5
        wd[rc.entry6] = sc.ui.entry6
        wd[rc.entry7] = sc.ui.entry7
        wd[rc.entry8] = sc.ui.entry8
        wd[rc.exit1] = sc.ui.exit1
        wd[rc.exit2] = sc.ui.exit2
        wd[rc.exit3] = sc.ui.exit3
        wd[rc.exit4] = sc.ui.exit4
        wd[rc.exit5] = sc.ui.exit5
        wd[rc.exit6] = sc.ui.exit6
        wd[rc.exit7] = sc.ui.exit7
        wd[rc.exit8] = sc.ui.exit8
        wd[rc.time1] = sc.ui.time1
        wd[rc.time2] = sc.ui.time2
        wd[rc.time3] = sc.ui.time3
        wd[rc.time4] = sc.ui.time4
        wd[rc.time5] = sc.ui.time5
        wd[rc.time6] = sc.ui.time6
        wd[rc.time7] = sc.ui.time7
        wd[rc.time8] = sc.ui.time8
        wd[rc.eshare1] = sc.ui.share1
        wd[rc.eshare2] = sc.ui.share2
        wd[rc.eshare3] = sc.ui.share3
        wd[rc.eshare4] = sc.ui.share4
        wd[rc.eshare5] = sc.ui.share5
        wd[rc.eshare6] = sc.ui.share6
        wd[rc.eshare7] = sc.ui.share7
        wd[rc.eshare8] = sc.ui.share8
        wd[rc.diff1] = sc.ui.diff1
        wd[rc.diff2] = sc.ui.diff2
        wd[rc.diff3] = sc.ui.diff3
        wd[rc.diff4] = sc.ui.diff4
        wd[rc.diff5] = sc.ui.diff5
        wd[rc.diff6] = sc.ui.diff6
        wd[rc.diff7] = sc.ui.diff7
        wd[rc.diff8] = sc.ui.diff8
        wd[rc.pl1] = sc.ui.pl1
        wd[rc.pl2] = sc.ui.pl2
        wd[rc.pl3] = sc.ui.pl3
        wd[rc.pl4] = sc.ui.pl4
        wd[rc.pl5] = sc.ui.pl5
        wd[rc.pl6] = sc.ui.pl6
        wd[rc.pl7] = sc.ui.pl7
        wd[rc.pl8] = sc.ui.pl8
        wd[rc.explain] = sc.ui.explain
        wd[rc.notes] = sc.ui.notes

        self.rc = rc
        self.wd = wd
        self.imageNames = None
        self.sc.loadLayoutForms(self)

    def save(self, wb, jf):
        '''
        Save wb as an excel file. If Permission Denied error is thrown, try renaming it.
        '''
        #Write the file
        jf.mkOutdir()
        saveName = jf.outpathfile
        count = 1
        saveName = saveName.replace('/', '\\')
        while True:
            try:
                wb.save(saveName)
            except PermissionError as ex:
                print()
                print(ex)
                print()
                print("Failed to create file {0}.{1}".format(saveName, ex))
                print("Images from the clipboard were saved  in {0}".format(jf.outdir))
                (nm, ext) = os.path.splitext(jf.outpathfile)
                saveName = "{0}({1}){2}".format(nm, count, ext)
                print("Will try to save as {0}".format(saveName))
                count = count+1
                if count == 6:
                    print("Giving up. PermissionError")
                    raise (PermissionError(
                        "Failed to create file {0}".format(saveName)))
                continue
            except Exception as ex:
                print(ex.__class__, ex)
            break

    def populateXLDailySummaryForm(self, mistake, ws, anchor):
        '''
        For the export to excel daily summary forms.
        Populate the daily Summary Form. The values are retrieved from self.ts via the
        DailyControl dialog object-without calling runDialog. The static labels are set earlier
        at the creation of the form shape/style. This method sets some statistics and notes for
        things like  average winners/losers etc.

        :params mistke:
        :params ws: The openpyxl Worksheet object
        :raise Value Error: When pl is misformatted and cannot be used.
        '''

        dc = DailyControl()
        dailySumData = dc.gatherDSumData(self.ts)
        anchor = (anchor[0], anchor[1] + mistake.numTrades + 5)
        for key in dailySumData.keys():
            rng = mistake.dailySummaryFields[key][0]
            if isinstance(rng, list):
                rng = rng[0]
            ws[tcell(rng, anchor=anchor)] = dailySumData[key]

    def populateXLMistakeForm(self, mistake, ws, imageLocation):
        '''
        For the export to excel the mistake summary form.
        Populate the dynamic parts of mistake summaries. That includes fomulas with references to
        tradeSummaries and hyperlinks to the same. The cell location for the links to tradeSummaries
        requires then anchor information in imageLocation and the specific cell within
        tradeSummaries found in mistakeFields. The return hyperlinks in the tradeSummaries forms are
        also translated here.

        :parmas mistake: A dataframe containing the info to populate the mistake summary form.
        :params ws: The openpyxl worksheet object.
        :parmas imageLocation: A list containing the locations in the worksheet for each of the
                               trades in tradeSummaries.
        '''

        # Populate the name fields as hyperlinks to tradeSummaries title cell and back.
        for i, (iloc, tradekey) in enumerate(zip(imageLocation, self.ts)):
            tsum = self.ts[tradekey]
            key = "name" + str(i+1)
            cell = mistake.mistakeFields[key][0][0]
            cell = tcell(cell, anchor=mistake.anchor)
            targetcell = (1, iloc[0])
            targetcell = tcell(targetcell)
            cellval = "{0} {1} {2}".format(i+1, tsum.Name.unique()[0], tsum.Account.unique()[0])
            link = "#{}!{}".format(ws.title, targetcell)

            ws[cell].hyperlink = (link)
            ws[cell] = cellval
            ws[cell].font = Font(color=colors.WHITE, underline="double")

            link = "#{}!{}".format(ws.title, cell)
            ws[targetcell].hyperlink = (link)
            ws[targetcell].font = Font(color=colors.WHITE, size=16, underline="double")

        # Populate the pl (loss) fields and the mistake fields. These are all formulas like =B31
        tokens = ["tpl", "pl", "mistake"]
        for token in tokens:
            for i in range(len(self.ts)):
                key = token + str(i+1)
                if isinstance(mistake.mistakeFields[key][0], list):
                    cell = mistake.mistakeFields[key][0][0]
                else:
                    cell = cell = mistake.mistakeFields[key][0]
                cell = tcell(cell, anchor=mistake.anchor)
            #     print(cell)
                formula = mistake.formulas[key][0]
                targetcell = mistake.formulas[key][1]
                targetcell = tcell(targetcell, anchor=(1, imageLocation[i][0]))
                formula = formula.format(targetcell)

                # print("ws[{0}]='{1}'".format(cell, formula))
                ws[cell] = formula

    def placeImages(self, imageLocation, ws, tf):
        '''
        '''

        # tradeSummaries = list()
        srf = SumReqFields()
        settings = QSettings('zero_substance', 'structjour')

        # response = askUser("Would you like to enter strategy names, targets and stops?   ")

        for loc in imageLocation:
            # tto = self.ts[tradekey]
            fname = loc[1][0]
            if not os.path.exists(fname):
                print(fname)
                outdir = settings.value('outdir')
                fname = os.path.join(outdir, fname)
                if not os.path.exists(fname):
                    continue
            img = Image(fname)

            if img:
                col = srf.maxcol() + 1
                col = tcell((col, 1))[0]
                cellname = col + str(loc[0])
                ws.add_image(img, cellname)

            #Put together the trade summary info for each trade and interview the trader
        
            #Place the format shapes/styles in the worksheet
            tf.formatTrade(ws, srf, anchor=(1, loc[0]))
    
    def populateXLDailyFormulas(self, imageLocation, ws):
        srf = SumReqFields()
        for loc, tradekey in zip(imageLocation, self.ts):
            tto = self.ts[tradekey]

            #populate the trade information
            for key in srf.tfcolumns:
                # print(key)
                # print(srf.tfcolumns[key][0])
                cell = srf.tfcolumns[key][0]
                if isinstance(cell, list):
                    cell = cell[0]
                tradeval = tto[key].unique()[0]

                # print ("{0:10} \t{3} \t{1:}\t{2} ".format(key, cell,
                # tradeval, tcell(cell, anchor=(1, loc[0]))))

                # Put some formulas in each trade Summary
                if key in srf.tfformulas:

                    anchor = (1, loc[0])
                    formula = srf.tfformulas[key][0]
                    args = []
                    for c in srf.tfformulas[key][1:]:
                        args.append(tcell(c, anchor=anchor))
                    tradeval = formula.format(*args)

                if not tradeval:
                    continue
                if isinstance(tradeval, (pd.Timestamp, dt.datetime, np.datetime64)):
                    tradeval = pd.Timestamp(tradeval)


                ws[tcell(cell, anchor=(1, loc[0]))] = tradeval

        return

    def imageDataFromQt(self, df, ldf):
        '''
        Gather the image names and determine the locations in the Excel doc to place them. Excel
        has a few things at top followed by trade summaries, charts and tables for each trade.
        Return with the image name/location data structure. The structure can be used for the Excel
        DataFrame-- to navigate summary form locations and just for the names
        :params df: The DataFrame representing the input file plus some stuff added in
                    processOutputFile
        :params ldf: A list of dataFrames. Each encapsulates a trade.
        :return (Imagelocation, df): ImageLocation contains information about the excel document
                    locations of trade summaries and image locations. The dataFrame df is the
                    outline used to create the workbook, ImageLocation will be used to stye it
                    and fill in the stuff.
        '''
        # Add rows and append each trade, leaving space for an image. Create a list of
        # names and row numbers to place images within the excel file (imageLocation
        # data structure).

        # Number of rows between trade summaries

        srf = SumReqFields()
        sumSize = srf.maxrow() + 5
        summarySize = sumSize
        spacing = 3
        frq = FinReqCol()
        newdf = DataFrameUtil.createDf(df, self.topMargin)

        df = newdf.append(df, ignore_index=True)

        imageLocation = list()
        count = 0
        imageNames = dict()
        for key in self.ts:
            images = []
            keys = self.ts[key].keys()
            charts = ['chart1', 'chart2', 'chart3']
            for chart in charts:
                if chart in keys:
                    images.append(self.ts[key][chart].unique()[0])
            key = key.split(' ')
            newkey = 'Trade ' + key[0]
            imageNames[newkey] = images

        for tdf in ldf:
            theKey = tdf[frq.tix].unique()[-1]
            imageName = imageNames[theKey]

            # Holds location, deprected name, image name, trade start time, trade duration as delta
            imageLocation.append([len(tdf) + len(df) + spacing,
                                  imageName,
                                  tdf.Start.unique()[-1],
                                  tdf.Duration.unique()[-1]])
            # print(count, imageName, len(imageLocation), len(tdf) + len(df) + 3)
            count = count + 1

            # Append the mini trade table then add rows to fit the tradeSummary form
            df = df.append(tdf, ignore_index=True)
            df = DataFrameUtil.addRows(df, summarySize)
        return imageLocation, df

    def exportExcel(self):
        '''
        Export to excel the trade tables, trade summaries, and daily forms
        '''

        # Create the space in dframe to add the summary information for each trade.
        # Then create the Workbook.
        settings = QSettings('zero_substance', 'structjour')
        val = settings.value('inputType')
        tu = DefineTrades(val)
        ldf = tu.getTradeList(self.df)

        ls = LayoutSheet(self.topMargin, len(self.df))
        # ldf = list(self.ts.values())
        imageLocation, dframe = self.imageDataFromQt(self.df, ldf)
        wb, ws, nt = ls.createWorkbook(dframe)

        tf = TradeFormat(wb)
        ls.styleTop(ws, len(nt.columns), tf)
        assert len(ldf) == len(imageLocation)

        mstkAnchor = (len(dframe.columns) + 2, 1)
        mistake = MistakeSummary(numTrades=len(ldf), anchor=mstkAnchor)
        mistake.mstkSumStyle(ws, tf, mstkAnchor)
        mistake.dailySumStyle(ws, tf, mstkAnchor)

        # tradeSummaries = ls.runSummaries(imageLocation, ldf, self.jf, ws, tf)
        self.placeImages(imageLocation, ws, tf)
        self.populateXLDailyFormulas(imageLocation, ws)
        

        self.populateXLMistakeForm(mistake, ws, imageLocation)
        self.populateXLDailySummaryForm(mistake, ws, mstkAnchor)


        self.save(wb, self.jf)
        print("Processing complete. Saved {}".format(self.jf.outpathfile))
        # return jf

    def getDF(self):
        return self.df

    def pickleitnow(self):
        name = f'.trades{self.jf.theDate.strftime(self.jf.dayformat)}.zst'
        fname = os.path.join(self.jf.outdir, name)
        with open(fname, "wb") as f:
            pickle.dump(self.df, f)
        settings = QSettings('zero_substance', 'structjour')
        settings.setValue('stored_trades', fname)

    def saveTheTradeObject(self, name):
        '''pickle tto list'''
        assert self.ts
        assert self.entries

        # if not self.df is None:
        #     self.reloadit()

        settings = QSettings('zero_substance', 'structjour')
        dfname = settings.value('stored_trades')
        df = None
        if dfname:
            with open(dfname, "rb") as f:
                df = pickle.load(f)
        if df is None and self.df:
            df = self.df
        if df is None:
            print('Failed to locate the trades information. Pickle FAILED')
        self.df = df

        with open(name, "wb") as f:
            pickle.dump((self.ts, self.entries, df), f)

    def loadSavedFile(self):
        '''
        Depickle a saved tto list. Clear then repopulate the tradeList widget. The first append
        will trigger the loading mechanism from tto to QT widgets
        '''
        name = self.sc.getSaveName()
        if not os.path.exists(name):
            print(f'Save file does not exist "{name}".')
            return None
        with open(name, "rb") as f:
            test = pickle.load(f)
            if len(test) == 2:
                print('Save is in the wrong format. Save and load it again to correct it')
                (self.ts, self.entries) = test
                # self.ts = test
            elif len(test) != 3:
                print('Something is wrong with this file')
                return
            else:
                (self.ts, self.entries, self.df) = test

        print('load up the trade names now')
        tradeSummaries = []
        self.sc.ui.tradeList.clear()
        for key in self.ts:
            self.sc.ui.tradeList.addItem(key)
            tradeSummaries.append(self.ts[key])
        try:
            self.sc.dControl.runDialog(self.df, self.ts)
        except AttributeError as e:
            print(e)

        # In prep to do the mistake summary and excel export, return the list it uses now
        # It might be good to use the dict self.ts instead
        return tradeSummaries

    def reloadit(self):
        from journal.statement import Statement_DAS as Ticket
        from journal.statement import Statement_IBActivity
        from journal.pandasutil import InputDataFrame

        infile = self.jf.inpathfile
        if not os.path.exists(infile):
            print("There is a problem. Unable to fully save this file.")
            return None
        if self.jf.inputType == 'IB_HTML':
            statement = Statement_IBActivity(self.jf)
            df = statement.getTrades_IBActivity(self.jf.inpathfile)
        elif  self.jf.inputType == 'DAS':
            tkt = Ticket(self.jf)
            df, self.jf = tkt.getTrades()
        # trades = pd.read_csv(self.jf.inpathfile)
        else:
            #Temporary
            print('Opening a non standard file name in DAS')
            tkt = Ticket(self.jf)
            df, self.jf = tkt.getTrades()

        idf = InputDataFrame()
        trades, success = idf.processInputFile(df, self.jf.theDate, self.jf)
        self.df = trades
        if not success:
            return

    def imageData(self, ldf):
        '''
        Create generic image names. Structjour will use this to create specific names that include
        interval info. Up to three images can be saved for each trade.
        '''

        frq = FinReqCol()
        imageNames = list()
        for tdf in ldf:
            dur = tdf[frq.dur].unique()[-1]
            if isinstance(dur, pd.Timedelta):
                dur = dur.__str__()
            dur = dur.replace(' ', '_')
            imageName = '{0}_{1}_{2}_{3}.{4}'.format(tdf[frq.tix].unique()[-1].replace(' ', ''),
                                                     tdf[frq.name].unique()[-1].replace(' ', '-'),
                                                     tdf[frq.start].unique()[-1], dur, 'png')
            imageName = imageName.replace(':', '')
            imageNames.append(imageName)
        return imageNames

    def runSummaries(self, ldf):
        '''
        This script creates the tto object for each trade in the input file and appends it to a
        list It also creates a generic name for assoiated images. That name will be altered for
        speific images that may be created via the stock api or added by the user. Finally the
        sript creates the tradeList key and adds it to the tradeList widget. The key is used to
        retrieve the tto data from the tradeList widget currentText selection.
        :params ldf: A list of DataFrames. Each df is a complete trade from initial purchace or
                    hold to 0 shares or hold.
        '''

        tradeSummaries = list()

        srf = SumReqFields()
        self.imageNames = self.imageData(ldf)
        assert len(ldf) == len(self.imageNames)
        self.sc.ui.tradeList.clear()
        for i, (imageName, tdf) in enumerate(zip(self.imageNames, ldf)):

            tto = TheTradeObject(tdf, False, srf)
            tto.runSummary(imageName)
            tradeSummaries.append(tto.TheTrade)
            # for key in self.wd.keys():
            #     print(key, tto.TheTrade[key].unique()[0])
            tkey = f'{i+1} {tto.TheTrade[srf.name].unique()[0]}'
            self.ts[tkey] = tto.TheTrade
            self.entries[tkey] = tto.entries
            self.sc.ui.tradeList.addItem(tkey)

        self.tradeSummaries = tradeSummaries
        return tradeSummaries

    def populateTradeSumForms(self, key):
        '''
        Use the widget dictionary (self.wd) and tto to populate the form. The images and related
        widgets are handled seperately.
        :Programming Note: Maybe handle the image together with everything else. All the relate
        widgets have entries and keys in tto. Cleaner. Easier to maintain. Back burner 4/27/19
        '''

        tto = self.ts[key]
        for wkey in self.wd:
            daVal = tto[wkey].unique()[0]
            if isinstance(daVal, (np.floating, float)):
                daVal = '{:.02f}'.format(daVal)
            elif isinstance(daVal, (np.integer, int)):
                daVal = '{}'.format(daVal)
            elif isinstance(daVal, (pd.Timestamp, dt.datetime, np.datetime64)):
                daVal = pd.Timestamp(daVal)
                daVal = daVal.strftime(self.timeFormat)
            elif wkey == "Strategy":
                continue
            self.wd[wkey].setText(daVal)
            # print(wkey)

        strat = tto['Strategy'].unique()[0]
        self.sc.loadStrategies(strat)

        self.sc.setChartTimes()
        iname1 = self.sc.ui.chart1Name.text()
        iname2 = self.sc.ui.chart2Name.text()
        iname3 = self.sc.ui.chart3Name.text()
        outdir = self.sc.getOutdir()
        ipathfilename1 = os.path.join(outdir, iname1)
        ipathfilename2 = os.path.join(outdir, iname2)
        ipathfilename3 = os.path.join(outdir, iname3)
        self.sc.loadImageFromFile(self.sc.ui.chart1, ipathfilename1)
        self.sc.loadImageFromFile(self.sc.ui.chart2, ipathfilename2)
        self.sc.loadImageFromFile(self.sc.ui.chart3, ipathfilename3)
        # print('never were here')

    def getEntries(self, key):
        '''
        The entries are pickled seperately in the dict self.entries. It uses parallel keys to
        self.ts. This data is trade information, read only and is used currently for chart
        generation. The data structure is:
        [price, time, share, pl, diff, entryOrExit]. Share is positive for buy, negative for sell.
        :params key: Trade name from the tradeList widget
        '''
        entries = self.entries[key]
        return entries

    def getChartData(self, key, ckey):
        '''
        Get the chart data from the tradeObject
        :params key: Trade name from the tradeList widget
        :params ckey: The widget name of the clickLabel will be one of 'chart1', 'chart2', or
                    'chart3'
        '''

        assert ckey in ('chart1', 'chart2', 'chart3')
        tto = self.ts[key]
        name = tto[ckey].unique()[0]
        begin = tto[ckey + 'Start'].unique()[0]
        end = tto[ckey + 'End'].unique()[0]
        if not isinstance(begin, (pd.Timestamp, dt.datetime, np.datetime64)) or (
                not isinstance(end, (pd.Timestamp, dt.datetime, np.datetime64))):
            print('WARNING: date type is not standard', type(begin))
            return None

        interval = tto[ckey + 'Interval'].unique()[0]

        return [name, begin, end, interval]

    def setChartData(self, key, data, ckey):
        '''
        Store the chart data in the trade object
        :params key: Trade name from the tradeList
        :params data: a list or list of lists. Each has: [start, end, interval, name]
        :params ckey: The widget name of the clickLabel will be one of 'chart1', 'chart2', or
                    'chart3'
        '''
        if self.ts:

            assert len(self.ts[key] == 1)
            if isinstance(ckey, list):
                for k, d in zip(ckey, data):
                    assert k in ['chart1', 'chart2', 'chart3']
                    self.ts[key].at[0, k] = d
                    return
            assert ckey in ['chart1', 'chart2', 'chart3']
            self.ts[key][ckey] = data[0]
            self.ts[key][ckey + 'Start'] = data[1]
            self.ts[key][ckey + 'End'] = data[2]
            self.ts[key][ckey + 'Interval'] = data[3]

    def reloadTimes(self, key):
        '''
        reload the time values for the trade time entries. This is done after toggling the date
        format to use (includes date info or not).
        '''
        tto = self.ts[key]
        twidgets = [self.sc.ui.time1, self.sc.ui.time2, self.sc.ui.time3, self.sc.ui.time4,
                    self.sc.ui.time5, self.sc.ui.time6, self.sc.ui.time7, self.sc.ui.time8]
        for i, widg in enumerate(twidgets):
            daVal = tto['Time' + str(i+1)].unique()[0]
            if isinstance(daVal, (pd.Timestamp, dt.datetime, np.datetime64)):
                daVal = pd.Timestamp(daVal)
                daVal = daVal.strftime(self.timeFormat)
            widg.setText(daVal)

        print(tto['Time1'], type(tto['Time1']))
        # self.sc.ui.time1

    def setTargVals(self, key, targ, diff, rr):
        '''Store the values affected by a change in the target value'''

        rc = self.rc
        tto = self.ts[key]
        tto[rc.targ] = targ
        tto[rc.targdiff] = diff
        if rr:
            tto[rc.rr] = rr

    def setStopVals(self, key, stop, diff, rr, maxloss):
        '''
        When the user enters a value in stoploss several things happen in the callback then
        here we store all the related values in tto. If the widgets have been marked clean,
        determine if a loss in PL exceeds maxloss. If so, save the data to tto
        '''

        rc = self.rc
        tto = self.ts[key]
        tto[rc.stoploss] = stop
        tto[rc.sldiff] = diff
        if rr:
            tto[rc.rr] = rr
        maxloss = 0.0 if not maxloss else maxloss
        tto[rc.maxloss] = maxloss

        lost = 0.0
        note = ''
        clean = tto['clean'].unique()[0]
        name = tto[rc.name].unique()[0]
        pl = tto[rc.pl].unique()[0]
        if maxloss and clean:
            if 'long' in name.lower() and diff >= 0:
                return lost, note, clean
            if 'short' in name.lower() and diff <= maxloss:
                return lost, note, clean
            assert maxloss < 0
            if maxloss > pl:
                lost = maxloss - pl
                tto[rc.mstkval] = lost
                note = 'Loss exceeds max loss!'
                tto[rc.mstknote] = note
        return (lost, note, clean)

    def setClean(self, key, b):
        '''
        Set set tto.clean to b
        :params key: The current trade name is the key found in the tradeList Combo box
        :params b: bool
        '''
        self.ts[key]['clean'] = b

    def setMstkVals(self, key, val, note):
        '''
        Set tto mstkval and mstknote to given values
        :params key: The current trade name is the key found in the tradeList Combo box
        :params val: Float, The value for the 'lost' widget
        :params note: The value for the sumNotes widget
        '''
        self.ts[key][self.rc.mstkval] = val
        self.ts[key][self.rc.mstknote] = note

    def setExplain(self, key, val):
        '''
        Set tto explain to val
        :params key: The current trade name is the key found in the tradeList Combo box
        :params val: The value for the explain widget
        '''
        self.ts[key][self.rc.explain] = val

    def setNotes(self, key, val):
        '''
        Set tto notes to val
        :params key: The current trade name is the key found in the tradeList Combo box
        :params val: The value for the notes widget
        '''
        self.ts[key][self.rc.notes] = val

    def setStrategy(self, key, val):
        '''Sets tto strategy to val'''
        self.ts[key][self.rc.strat] = val

    def getStrategy(self, key):
        val = self.ts[key][self.rc.strat].unique()[0]
        return val

    def getImageName(self, key, wloc, uinfo=''):
        '''
        This is an image name for a pasted image. We don't know the interval or time limits beyond
        the trade times. Remove any interval info, add the widget name (e.g. 'chart1') to the
        generic name.
        :params key: The name of the trade from the tradeList
        :params wloc: The name of the clickLabel widget
        :params uinfo: Misc string to add to the name
        '''

        name = self.ts[key][wloc].unique()[0]
        data = self.getChartData(key, wloc)
        if name:
            n, ext = os.path.splitext(name)
            if n.endswith('min'):
                n = n[:-5]
            wwloc = wloc
            if n.find(wwloc) > 0:
                wwloc = ''
            if uinfo and n.find(uinfo) > 0:
                uinfo = ''
            n = n + wwloc + uinfo + ext
        elif data:
            # Just in case the name is missing
            b = data[1]
            e = data[2]

            begin = qtime2pd(b)
            end = qtime2pd(e)
            # begin = pd.Timestamp(b.date().year(), b.date().month(), b.date().day(),
            #                      b.time().hour(), b.time().minute(), b.time().second())
            # end = pd.Timestamp(e.date().year(), e.date().month(), e.date().day(),
            #                    e.time().hour(), e.time().minute(), e.time().second())
            delt = end - begin
            bstring = begin.strftime('%H%M%S')
            estring = delt.__str__().replace(':', '.')
            n = 'Trade{}_{}_{}_{}_{}.png'.format(
                key, bstring, estring, wloc, uinfo)
            n = n.replace(' ', '_')
        if data[0] != n:
            data[0] = n
            self.setChartData(key, data, wloc)
        return n

    def getImageNameX(self, key, wloc):
        '''
        This is a chart name for which we know the candle interval because we gave that parameter
        to the stock api.
        '''
        name = key.replace(' ', '_')
        data = self.getChartData(key, wloc)
        if data:
            # Just in case the name is missing
            b = pd.Timestamp(data[1])
            e = pd.Timestamp(data[2])

            delt = e - b
            bstring = b.strftime('%H%M%S')
            estring = delt.__str__().replace(':', '.')
            istring = str(data[3]) + 'min'
            n = 'Trade{}_{}_{}_{}.png'.format(
                name, bstring, estring, istring)
            n = n.replace(' ', '_')
        if data[0] != n:
            data[0] = n
            self.setChartData(key, data, wloc)
        return n
