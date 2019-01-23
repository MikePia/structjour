from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (PatternFill,  Border, Side, 
                             Alignment, Font, NamedStyle)

from withstyle.thetradeobject import SumReqFields

#global variable
srf = SumReqFields()

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.
    I am using this borrowed code for assigning borders in a merged 
    range

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
#         ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def c(icell, end=None, anchor= None) :
    '''
    Translates numerical coordinates as tuplets to excels  string addresses e.g., (1,1) returns 'A1' 
    :params:icell: The cell to translate.
    :parmas:end: Creates an excel range with icell at the top left and end at the bottom right.
    :params:anchor: A translation amount. The cell or range  will be translated by the amount of the cell -1
                    to account for all cells beginning at (1,1) not(0,0)
    '''
    if anchor :
        anchor = (anchor[0] - 1, anchor[1] - 1)
        icell = (icell[0] + anchor[0], icell[1] + anchor[1])
        if end :
            end = (end [0] + anchor[0], end[1] + anchor[1])

    # mini recursion to handle range
    if end is not None:
        return "{0}:{1}".format(c(icell),c(end))
    # For the single coordinate
    return "{0}{1}".format(get_column_letter(icell[0]),icell[1])
    

class TradeFormat (object):
    ''' 
    Deal with everything that formats the openpyxl object to create the excel doc. That includes:
        Create the shape of the trade summary in an excel file.
        Apply the styles defined here and named in for each cell in SumReqFields.tfcolumns
        Create the sheet
        '''
    # TODO openpyxl throws an exception if you name a named style twice.  Do something to deal with it. Either catch the exception or check if the style exists or 
    # make this a singleton (---no singleton, it doesn't address the issue)
    def __init__(self, wb, a=(1,1) ):
        self.tradeAnchor = a
        
        self.styles = dict()
        
            
            
        # This code is subject to change on a whim. Primary importance is easy to find styles and 
        # consistent layout
        #######################################         titleStyle       ###################################################
        titleStyle = NamedStyle(name="titleStyle")
        titleStyle.font = Font(color="FFFFFF", size=16)
        titleStyle.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        titleStyle.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        titleStyle.border = Border(left=Side(style='double'),
                                 right=Side(style='double'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        self.addNamedStyle(titleStyle, 'titleStyle', wb)
        
                
        #######################################         titleLeft       ###################################################
        titleLeft= NamedStyle(name="titleLeft")
        titleLeft.font = Font(color="FFFFFF", size=16)
        titleLeft.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        titleLeft.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        titleLeft.border = Border(left=Side(style='double'),
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        self.addNamedStyle(titleLeft, 'titleLeft', wb)
        
        #######################################     titleNumberRight     ###################################################
        titleNumberRight= NamedStyle(name="titleNumberRight")
        titleNumberRight.font = Font(color="FFFFFF", size=16)
        titleNumberRight.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        titleNumberRight.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        titleNumberRight.border = Border(right=Side(style='double'),
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        titleNumberRight.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        self.addNamedStyle(titleNumberRight, 'titleNumberRight', wb)

        #######################################     finalNoteStyle     ###################################################
        finalNoteStyle = NamedStyle(name = "finalNoteStyle")
        finalNoteStyle.font = Font(color="FFFFFF", size=8)
        finalNoteStyle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        finalNoteStyle.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        finalNoteStyle.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        
        self.addNamedStyle(finalNoteStyle, 'finalNoteStyle', wb)

        #######################################     titleRight     ###################################################
        titleRight= NamedStyle(name="titleRight")
        titleRight.font = Font(color="FFFFFF", size=16)
        titleRight.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        titleRight.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        titleRight.border = Border(right=Side(style='double'),
                           top=Side(style='double'), 
                           bottom=Side(style='double'))
#         titleRight.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(titleRight, 'titleRight', wb)

        #######################################     normStyle     ###################################################        
        normStyle = NamedStyle(name = "normStyle")
        normStyle.font = Font(color="FFFFFF", size=11)
        normStyle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        normStyle.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normStyle.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        normStyle.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normStyle, 'normStyle', wb)
       
        #######################################     normalNumber     ###################################################        
        normalNumber= NamedStyle(name = "normalNumber")
        normalNumber.font = Font(color="FFFFFF", size=11)
        normalNumber.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumber.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumber.border = Border(left=Side(style='double'),
                                 right=Side(style='double'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        normalNumber.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumber, 'normalNumber', wb)
        
        
        
        #######################################     normalFraction     ###################################################
        normalFraction = NamedStyle(name = "normalFraction")
        normalFraction.font = Font(color="FFFFFF", size=11)
        normalFraction.alignment = Alignment(horizontal="left", vertical="bottom")
        normalFraction.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalFraction.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        normalFraction.number_format = '# ??/??'
        
        self.addNamedStyle(normalFraction, 'normalFraction',  wb) 

        #######################################     linkStyle    ###################################################        
        linkStyle = NamedStyle(name="linkStyle")
        linkStyle.font = Font(color="FFFFFF", size=16)
        linkStyle.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        linkStyle.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        linkStyle.border = Border(left=Side(style='double'),
                                 right=Side(style='double'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='double'))
        self.addNamedStyle(linkStyle, 'linkStyle', wb)
        
        #######################################     normalNumberTopLeft     ###################################################        
        normalNumberTopLeft= NamedStyle(name = "normalNumberTopLeft")
        normalNumberTopLeft.font = Font(color="FFFFFF", size=11)
        normalNumberTopLeft.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumberTopLeft.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumberTopLeft.border = Border(left=Side(style='double'),
                                 right=Side(style='thin'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='thin'))
        normalNumberTopLeft.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumberTopLeft, 'normalNumberTopLeft', wb)

        
        #######################################     normalNumberTop     #################################################        
        normalNumberTop= NamedStyle(name = "normalNumberTop")
        normalNumberTop.font = Font(color="FFFFFF", size=11)
        normalNumberTop.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumberTop.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumberTop.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='thin'))
        normalNumberTop.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumberTop, 'normalNumberTop', wb)       
        
        #######################################     normalNumberTopRight     #################################################        
        normalNumberTopRight = NamedStyle(name = "normalNumberTopRight")
        normalNumberTopRight.font = Font(color="FFFFFF", size=11)
        normalNumberTopRight.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumberTopRight.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumberTopRight.border = Border(left=Side(style='thin'),
                                 right=Side(style='double'), 
                                 top=Side(style='double'), 
                                 bottom=Side(style='thin'))
        normalNumberTopRight.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumberTopRight, 'normalNumberTopRight', wb)
        
        #######################################     normalNumberLeft     #################################################        
        normalNumberLeft= NamedStyle(name = "normalNumberLeft")
        normalNumberLeft.font = Font(color="FFFFFF", size=11)
        normalNumberLeft.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumberLeft.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumberLeft.border = Border(left=Side(style='double'),
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
        normalNumberLeft.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumberLeft, 'normalNumberLeft', wb)    
        
        #######################################     normalNumberRight     #################################################        
        normalNumberRight= NamedStyle(name = "normalNumberRight")
        normalNumberRight.font = Font(color="FFFFFF", size=11)
        normalNumberRight.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumberRight.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumberRight.border = Border(left=Side(style='thin'),
                                 right=Side(style='double'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
        normalNumberRight.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumberRight, 'normalNumberRight', wb)    
        
        #######################################     normalNumberInside     #################################################        
        normalNumberInside= NamedStyle(name = "normalNumberInside")
        normalNumberInside.font = Font(color="FFFFFF", size=11)
        normalNumberInside.alignment = Alignment(horizontal="left", vertical="bottom")
        normalNumberInside.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalNumberInside.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
        normalNumberInside.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalNumberInside, 'normalNumberInside', wb)    
        
        #######################################     normalSubLeft     #################################################
        normalSubLeft= NamedStyle(name = "normalSubLeft")
        normalSubLeft.font = Font(color="FABF8F", size=11)
        normalSubLeft.alignment = Alignment(horizontal="left", vertical="bottom")
        normalSubLeft.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalSubLeft.border = Border(left=Side(style='double'),
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

        self.addNamedStyle(normalSubLeft, 'normalSubLeft', wb)
       
        
        #######################################     normalSub     #################################################
        normalSub= NamedStyle(name = "normalSub")
        normalSub.font = Font(color="FABF8F", size=11)
        normalSub.alignment = Alignment(horizontal="left", vertical="bottom")
        normalSub.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalSub.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

        self.addNamedStyle(normalSub,  'normalSub', wb)
        
        #######################################     normalSubRight     #################################################
        normalSubRight= NamedStyle(name = "normalSubRight")
        normalSubRight.font = Font(color="FABF8F", size=11)
        normalSubRight.alignment = Alignment(horizontal="left", vertical="bottom")
        normalSubRight.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalSubRight.border = Border(left=Side(style='thin'),
                                 right=Side(style='double'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
#         normalSubLeft.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalSubRight, 'normalSubRight', wb)
        
        #######################################     normalSubNumberBottomLeft     #################################################
        normalSubNumberBottomLeft= NamedStyle(name = "normalSubNumberBottomLeft")
        normalSubNumberBottomLeft.font = Font(color="FABF8F", size=11)
        normalSubNumberBottomLeft.alignment = Alignment(horizontal="left", vertical="bottom")
        normalSubNumberBottomLeft.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalSubNumberBottomLeft.border = Border(left=Side(style='double'),
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='double'))
        normalSubNumberBottomLeft.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalSubNumberBottomLeft, 'normalSubNumberBottomLeft', wb)
        
        #######################################     normalSubNumberBottom     #################################################
        normalSubNumberBottom= NamedStyle(name = "normalNumberBottom")
        normalSubNumberBottom.font = Font(color="FABF8F", size=11)
        normalSubNumberBottom.alignment = Alignment(horizontal="left", vertical="bottom")
        normalSubNumberBottom.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalSubNumberBottom.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='double'))
        normalSubNumberBottom.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalSubNumberBottom, 'normalSubNumberBottom', wb)
        
        #######################################     normalSubNumberBottomRight     #################################################
        normalSubNumberBottomRight= NamedStyle(name = "normalSubNumberBottomRight")
        normalSubNumberBottomRight.font = Font(color="FABF8F", size=11)
        normalSubNumberBottomRight.alignment = Alignment(horizontal="left", vertical="bottom")
        normalSubNumberBottomRight.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        normalSubNumberBottomRight.border = Border(left=Side(style='thin'),
                                 right=Side(style='double'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='double'))
        normalSubNumberBottomRight.number_format = '"$"#,##0.00_);[Red]\("$"#,##0.00\)'
        
        self.addNamedStyle(normalSubNumberBottomRight, 'normalSubNumberBottomRight', wb)
        
        #######################################     explain     #################################################
        explain = NamedStyle(name="explain")
        explain.font = Font(color="000000", size=10)
        explain.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
        explain.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        explain.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        self.addNamedStyle(explain, 'explain', wb)
        
        #######################################     noteStyle     #################################################
        noteStyle = NamedStyle(name="noteStyle")
        noteStyle.font = Font(color="E26B0A", size=10)
        noteStyle.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
        noteStyle.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        noteStyle.border = Border(left=Side(style='double'),
                                  right=Side(style='double'), 
                                  top=Side(style='double'), 
                                  bottom=Side(style='double'))
        self.addNamedStyle(noteStyle, 'noteStyle', wb)
        
    def addNamedStyle(self, ns, strNs, wb):
        '''
        Register named style if it is not lready and add the style to the dict self.styles
        :params:ns The NamedStyle
        :params:strNs: The NamedStyle represented as a string
        :params:wb Workbook
        '''
        if strNs not in wb.named_styles :
            wb.add_named_style(ns)
            
        self.styles[strNs] = ns
        

    def mergeStuff(self, ws, begin, end, anchor=None) :
        '''
        Perform a merge cell
        :params:ws: The Worksheet 
        :params:begin: The top left cell as a tuple of ints (2,3)
        :params:end: The bottom right cell as a tuple of ints
        :parmas:anchor: Translation
        '''
        if anchor :
            anchor = (anchor[0] - 1, anchor[1] - 1)
            begin = (begin[0] + anchor[0], begin[1] + anchor[1])
            end = (end [0] + anchor[0], end[1] + anchor[1])

        ws.merge_cells(start_row=begin[1], 
                       start_column=begin[0], 
                       end_row=end[1], 
                       end_column=end[0])

        return (begin, end)
    

    def popTheTrade(self, ws, tf, row):
        anchor = "A:" + row
        srf.tfcolumns
        
        
    def formatTrade(self, ws, tstyle=None, anchor = None) :
        if anchor:
            self.tradeAnchor = anchor
        anc = self.tradeAnchor
        
        for val in  srf.tfcolumns.values() :
            if isinstance(val[0], list) :
                self.mergeStuff(ws, val[0][0], val[0][1], anchor=anc)
                rng = c(val[0][0], val[0][1], anchor=anc)
                
                style_range(ws, rng, self.styles[val[1]].border)
                ws[c(val[0][0], anchor=anc)].style = self.styles[val[1]]

            else :
                ws[c(val[0], anchor=anc)].style = self.styles[val[1]]
        
            
#         ws[c((1,11), anchor=anc)] = "(Technical description of the trade.)"
#         ws[c((1,17), anchor=anc)] = "(Evaluation of the trade)"
    
    


  
# 
#     def c(self, icell,  add= None) :
#         if add :
#             icell = (icell[0] + add[0], icell[1] + add[1])
#         return "{0}{1}".format(get_column_letter(icell[0]),icell[1])
    
   
    
   

# 
# wb=Workbook()
# 
# ws = wb.active
# tf = TradeFormat(wb, a = ((1,45)))
# tf.formatTrade(ws)
# tf.formatTrade(ws, anchor=(7,96))
# 
# 
# wb.save('out/justlookin.xlsx')
# print("done!")












