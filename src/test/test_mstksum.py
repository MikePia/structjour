'''
Test the methods in the module journal.mstksum

@created_on Feb 20, 2019

@author: Mike Petersen
'''
import os
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from journal.mstksum import MistakeSummary
from journal.tradestyle import TradeFormat
from journal.tradestyle import c as tcell

# pylint: disable = C0103




class TestMistakeSummary(TestCase):
    '''
    Run all of structjour with a collection of input files and test the outcome
    '''

    def __init__(self, *args, **kwargs):
        super(TestMistakeSummary, self).__init__(*args, **kwargs)

        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

    def test_mstkSumStyles(self):
        '''
        Test the method MistakeSummary.mstkSumStyle.
        '''
        wb = Workbook()
        ws = wb.active
        tf = TradeFormat(wb)

        ms = MistakeSummary(5)

        ms.mstkSumStyle(ws, tf)

        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        # test that each listed item in ms.mistakeFields has a corresponding style set in the
        # appropriate cell in the re-opened workbook.
        for x in ms.mistakeFields:
            entry = ms.mistakeFields[x]
            if isinstance(entry[0], list):
                cell = tcell(entry[0][0])
            else:
                cell = tcell(entry[0])

            # print(cell, entry[1], ws[cell].style)
            self.assertEqual(entry[1], ws[cell].style)

        headers = ['title', 'headname', 'headpl', 'headLossPL', 'headmistake']
        for h in headers:
            if isinstance(ms.mistakeFields[h][0], list):
                cell = tcell(ms.mistakeFields[h][0][0])
            else:
                cell = tcell(ms.mistakeFields[h][0])
            self.assertEqual(ws2[cell].value, ms.mistakeFields[h][2])
            # print(cell, ws2[cell].value, '----> <----', ms.mistakeFields[h][2])


def notmain():
    '''Run some local code'''
    t = TestMistakeSummary()
    t.test_mstkSumStyles()


if __name__ == '__main__':
    notmain()
