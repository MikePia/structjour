'''
Test the methods and functions in journal.tradestyle
@created_on Feb 12, 2019

@author: Mike Petersen
'''
import os
from unittest import TestCase

from openpyxl import load_workbook
from openpyxl import Workbook

from journal.tradestyle import TradeFormat, c
from journal.thetradeobject import SumReqFields


# pylint: disable = C0103, W0613, W0603


class TestTradeFormat(TestCase):
    '''
    Test the methods in journal.tradestyle.TradeFormat and also the functions in the module.
    '''

    def __init__(self, *args, **kwargs):
        super(TestTradeFormat, self).__init__(*args, **kwargs)

        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

    def test_TradeFormatAddNamedStyle(self):
        '''
        Test the object formation and the creation of the styles in __init__ and addNamedStyle for
        the Workbook. Specifically test that the named style is created in the workbook and is
        still there when you close it and open it.
        '''

        wb = Workbook()
        ws = wb.active
        tf = TradeFormat(wb)

        # Use a list to verify the order
        styleList = list(tf.styles.keys())
        if not os.path.exists("out/"):
            os.mkdir("out/")

        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        # Write one cell for each named style in TradeFormat
        for i, key in enumerate(styleList):
            #     print(key, c((1, i+1)))
            ws[c((1, i+1))] = key
            ws[c((1, i+1))].style = tf.styles[key]
        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        # Open it back up and check the the named styles are where we think they are
        for i, key in enumerate(styleList):
            #     print(key,ws2[c((1, i+1))].style )
            self.assertEqual(key, ws2[c((1, i+1))].style)
        # print('Done')

    def test_mergeStuff(self):
        '''
        Test the method tradestyle.TradeFormat.mergeStuff. Specificaly test that a merged cell
        group exists where it is should be after saving and opening an xlsx file.
        '''
        wb = Workbook()
        ws = wb.active

        begin = (1, 1)
        end = (5, 5)
        anchor = (3, 3)
        t = TradeFormat(wb)

        b, e = t.mergeStuff(ws, begin=begin, end=end, anchor=anchor)

        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)
        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        x = ws2.merged_cells.ranges

        # print(b, (x[0].min_col, x[0].min_row))
        # print(e, (x[0].max_col, x[0].max_row))
        self.assertEqual(b, (x[0].min_col, x[0].min_row))
        self.assertEqual(e, (x[0].max_col, x[0].max_row))
        os.remove(dispath)

    def test_formatTrade(self):
        '''
        Test the method TradeFormat.formatTrade. Specifically test that each of the elements in
        srf.tfcolumns has a corresponding elementcorrectly styled in the re-opened workbook and
        that each merged element in srf.tfcolumns is found in the the worsheet.
        '''
        wb = Workbook()
        ws = wb.active

        t = TradeFormat(wb)
        srf = SumReqFields()
        ws = wb.active

        t.formatTrade(ws, srf)

        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)
        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        # test that each listed item in srf.tfcolumns has a corresponding style set in the
        # appropriate cell of the re-opened workbook
        for x in srf.tfcolumns:
            address = srf.tfcolumns[x][0]
            st = srf.tfcolumns[x][1]
            if isinstance(address, list):
                assert ws2[c(address[0])].style == st
            else:
                assert ws2[c(address)].style == st

        # test that each list element has a corresponding merged cell group in the worbook
        listofmerge = [c(srf.tfcolumns[x][0][0], srf.tfcolumns[x][0][1])  for x in srf.tfcolumns if isinstance(srf.tfcolumns[x][0], list)]
        wsmerged = ws.merged_cells.ranges
        self.assertEqual(len(listofmerge), len(wsmerged))
        for msmerge in wsmerged:
            self.assertTrue(str(msmerge) in listofmerge)


    def test_c(self):
        '''Test the function c in the module tradestyle'''
        self.assertEqual(c((3, 4)), 'C4')
        self.assertEqual(c((1, 1), anchor=(4, 5)), 'D5')

        self.assertEqual(c((2, 3), (3, 7)), 'B3:C7')

        self.assertEqual(c((2, 3), (3, 7), (2, 2)), 'C4:D8')

def notmain():
    '''Run some local code'''
    t = TestTradeFormat()
    # t.test_TradeFormatAddNamedStyle()
    # t.test_mergeStuff()
    # t.test_formatTrade()
    t.test_c()


if __name__ == '__main__':
    notmain()
