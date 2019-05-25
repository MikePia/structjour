# Structjour -- a daily trade review helper
# Copyright (C) 2019 Zero Substance Trading
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.

'''

Test the methods in the module journal.dailysumforms

@created_on Feb 20, 2019

@author: Mike Petersen
'''
import os
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from journal.dailysumforms import MistakeSummary
from journal.tradestyle import TradeFormat
from journal.tradestyle import c as tcell

# pylint: disable = C0103




class TestMistakeSummary(TestCase):
    '''
    Run all of structjour with a collection of input files and test the outcome
    '''

    def __init__(self, *args, **kwargs):
        super(TestMistakeSummary, self).__init__(*args, **kwargs)

        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

    def test_mstkSumStyles(self):
        '''
        Test the method MistakeSummary.mstkSumStyle.
        '''
        wb = Workbook()
        ws = wb.active
        tf = TradeFormat(wb)

        ms = MistakeSummary(5)

        ms.mstkSumStyle(ws, tf)

        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        # test that each listed item in ms.mistakeFields has a corresponding style set in the
        # appropriate cell in the re-opened workbook.
        for x in ms.mistakeFields:
            cell = ''
            entry = ms.mistakeFields[x]
            if isinstance(entry[0], list):
                cell = tcell(entry[0][0])
            else:
                cell = tcell(entry[0])

            # print(cell, entry[1], ws[cell].style)
            self.assertEqual(entry[1], ws[cell].style)

            if len(entry) == 3:
                self.assertEqual(ws2[cell].value, ms.mistakeFields[x][2])
        os.remove(dispath)

    def test_dailySumStyle(self):
        '''
        Test the method MistakeSummary.mstkSumStyle.
        '''
        wb = Workbook()
        ws = wb.active
        tf = TradeFormat(wb)

        numTrades = 5

        ms = MistakeSummary(numTrades)

        ms.dailySumStyle(ws, tf)

        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        wb.save(dispath)

        anchor = (1, ms.anchor[1] + ms.numTrades + 5)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        # test that each listed item in ms.mistakeFields has a corresponding style set in the
        # appropriate cell in the re-opened workbook.
        for x in ms.dailySummaryFields:
            cell = ''
            entry = ms.dailySummaryFields[x]
            if isinstance(entry[0], list):
                cell = tcell(entry[0][0], anchor=anchor)
            else:
                cell = tcell(entry[0], anchor=anchor)
            self.assertEqual(entry[1], ws2[cell].style)

            # Get the entries with the static values-- the headers
            if len(entry) == 3:
                self.assertEqual(ws2[cell].value, ms.dailySummaryFields[x][2])
        os.remove(dispath)


def notmain():
    '''Run some local code'''
    t = TestMistakeSummary()
    t.test_mstkSumStyles()
    # t.test_dailySumStyle()


if __name__ == '__main__':
    notmain()
