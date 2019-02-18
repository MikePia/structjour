'''
Test the methods in the module layoutsheet

@created_on Feb 8, 2019

@author: Mike Petersen
'''
import os
from random import randint
from unittest import TestCase
from unittest.mock import patch
from collections import deque

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

from journalfiles import JournalFiles
from journal.pandasutil import InputDataFrame
from journal.tradeutil import TradeUtil
from journal.layoutsheet import LayoutSheet
from journal.tradestyle import TradeFormat, c
# pylint: disable = C0103, W0613, W0603, W0212


D = deque()


def mock_askUser(dummy, dummy2):
    '''
    Mock the specific askUser function that asks how many shares are currently owned or owned
    before trading today.
    '''
    global D
    x = D.popleft()
    # print("Returning from the mock ", x)
    return x


class TestLayoutSheet(TestCase):
    '''
    Run all of structjour with a collection of input files and test the outcome
    '''

    def __init__(self, *args, **kwargs):
        super(TestLayoutSheet, self).__init__(*args, **kwargs)

        ddiirr = os.path.dirname(__file__)
        os.chdir(os.path.realpath(ddiirr + '/../'))

        self.dadata = deque(
            [[-4000], [3923, -750], [], [600, 241, 50], [-169],
             [], [0, -600], [], [0, 750, 50], [-600], ])

        # Input test files can be added here. And place the test data in testdata.xlsx. Should add
        # files with potential difficulties
        self.infiles = ['trades.1116_messedUpTradeSummary10.csv', 'trades.8.WithHolds.csv',
                        'trades.8.csv', 'trades.907.WithChangingHolds.csv',
                        'trades_190117_HoldError.csv', 'trades.8.ExcelEdited.csv',
                        'trades.910.tickets.csv', 'trades_tuesday_1121_DivBy0_bug.csv',
                        'trades.8.WithBothHolds.csv', 'trades1105HoldShortEnd.csv']

        # self.tests = self.getTestData(r'C:\python\E\structjour\src\data')

    def getTestData(self, indir):
        '''
        Open the csv file testdata and oraganze the data into a usable data structure. The file is
        necessarily populated by 'hand.' To add a file to test, copy it to the data dir and enter
        the information to test.
        :return data: List  containing the data to check against the output files.
        '''

        df = pd.read_excel(os.path.join(indir, 'testdata.xlsx'))

        l = len(df)
        # print(l)
        data = list()

        data = list()
        for i, row in df.iterrows():
            entry = list()
            if not pd.isnull(df.at[i, 'Order']):
                entry.extend((row['Order'], row['NumTrades'], row['Name']))
                j = i
                trades = list()
                begginning = True
                while j < l and not pd.isnull(df.at[j, 'Ticker']):
                    # Check these specific trades
                    if not begginning:
                        if isinstance(df.at[j, 'Name'], str):
                            break

                    trades.append([df.at[j, 'Ticker'], df.at[j, 'Account'],
                                   df.at[j, 'Side'], df.at[j, 'Held'], df.at[j, 'Pos'], ])
                    begginning = False
                    j = j+1
                entry.append(trades)
                data.append(entry)
        return data

    @patch('journal.pandasutil.askUser', side_effect=mock_askUser)
    def test_createImageLocation(self, unusedstub):
        '''Run structjour'''
        #Assert the initial entry location in imageLocation is figured by these factors
         # Assert all entries are figured by summarySize and len of the minittrade
        # Assert the minitable locations are offset by length of the minitrade tables +
                # ls.spacing, and the first entry in the Leftmost column starts with 'Trade'
                 # Assert third entry of imageLocation begins with 'Trade' and contains
                #       ['Long', 'Short']
                # Assert the 4th entry is a timestamp
                # Assert the 5th entry is a time delta
        global D


        for tdata, infile in zip(self.dadata, self.infiles):
            # :::::::::  Setup   ::::::::
            D = deque(tdata)
            # infile = 'trades.8.csv'
            indir = 'data/'
            mydevel = False
            jf = JournalFiles(indir=indir, infile=infile, mydevel=mydevel)

            trades = pd.read_csv(jf.inpathfile)

            idf = InputDataFrame()
            trades = idf.processInputFile(trades)

            tu = TradeUtil()
            inputlen, dframe, ldf = tu.processOutputDframe(trades)
            # ::::::::::: end setup :::::::::::::

            sumSize = 25
            margin = 25
            spacing = 3
            ls = LayoutSheet(sumSize, margin, inputlen, spacing=spacing)
            imageLocation, dframe = ls.createImageLocation(dframe, ldf)

            import datetime as dt

            for count, t in enumerate(imageLocation):
                if count == 0:
                    initialEntry = ls.inputlen + ls.topMargin + ls.spacing + len(ldf[0]) + 2
                    self.assertEqual(t[0], initialEntry)

                else:
                    nextloc = imageLocation[count-1][0] + len(ldf[count]) + ls.summarySize
                    self.assertEqual(t[0], nextloc)

                t_entry = t[0] - (spacing + len(ldf[count]))
                self.assertTrue(dframe.iloc[t_entry][0].startswith('Trade'))
                self.assertEqual(len(dframe.iloc[t_entry-1][0]), 0)
                self.assertTrue(t[2].startswith('Trade'))
                self.assertTrue(t[2].find('Long') > 0 or t[2].find('Short') > 0)
                self.assertTrue(isinstance(pd.Timestamp('2019-11-11 ' + t[3]), dt.datetime))
                self.assertTrue(isinstance(t[4], dt.timedelta))

            print('Done test_createImageLocation', infile)

    def test_createWorkbook(self):
        '''
        Test the method journal.layoutsheet.LayoutSheet.createWorkbook
        '''

        df = pd.DataFrame(np.random.randint(0, 100, size=(100, 7)), columns=list('ABCDEFG'))
        # df
        sumSize = 25
        margin = 25
        spacing = 3
        inputlen = len(df)
        ls = LayoutSheet(sumSize, margin, inputlen, spacing=spacing)

        wb, ws, df = ls.createWorkbook(df)


        for row, (i, dfrow) in zip(ws, df.iterrows()):
            # We inserted the column headers in this row (ws starts with 1, not 0)
            if i + 1 == ls.topMargin:
                for ms, x in zip(row, df.columns):
                    print(x, ms.value)
                    self.assertEqual(x, ms.value)
            # everything else is verbatim
            else:
                for ms, x in zip(row, dfrow):
                    # print(x, ms.value)
                    self.assertEqual(x, ms.value)


        wb.save("out/SCHNOrK.xlsx")

    def test_styleTop(self):
        '''
        Test the method layoutsheet.LayoutSheet.styleTop. We necessarily know too much about it.
        For example we know the merge sizes and the styles that the method hard codes. Note that
        we are using a protected member of Worksheet ws._tables
        '''
        quoteRange = [(1, 1), (13, 5)]
        noteRange = [(1, 6), (13, 24)]
        quoteStyle = 'normStyle'
        noteStyle = 'explain'
        sumSize = 25
        margin = 25
        inputlen = 50   #len(df)

        wb = Workbook()
        ws = wb.active
        tf = TradeFormat(wb)

        # Make sure the out dir exists
        if not os.path.exists("out/"):
            os.mkdir("out/")

        # Make sure the file we are about to create does not exist
        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        # Create table header and data in the ws
        headers = ['Its', 'the', 'end', 'of', 'the', 'world', 'as', 'we',
                   'know', 'it.', 'Bout', 'Fn', 'Time!']
        for i in range(1, 14):
            ws[c((i, 25))] = headers[i-1]

        ls = LayoutSheet(sumSize, margin, inputlen)
        for x in range(ls.topMargin+1, ls.inputlen + ls.topMargin+1):
            for xx in range(1, 14):
                ws[c((xx, x))] = randint(-1000, 10000)


        ls.styleTop(ws, 13, tf)

        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        x = ws2.merged_cells.ranges
        listOfMerged = list()
        listOfMerged.append(c((quoteRange[0])) + ':' +  c((quoteRange[1])))
        listOfMerged.append(c((noteRange[0])) + ':' +  c((noteRange[1])))
        for xx in x:
            # print (str(xx) in listOfMerged)
            self.assertTrue(str(xx) in listOfMerged)
        self.assertEqual(ws[c(quoteRange[0])].style, quoteStyle)
        self.assertEqual(ws[c(noteRange[0])].style, noteStyle)

        self.assertEqual(len(ws._tables), 1)


        begin = c((1, ls.topMargin))

        end = c((13, ls.topMargin + ls.inputlen))
        tabRange = f'{begin}:{end}'
        self.assertEqual(tabRange, ws._tables[0].ref)

        os.remove(dispath)

    def test_styleTopwithnothin(self):
        '''
        Test the method layoutsheet.LayoutSheet.styleTop. Test that it still works without
        table data. We still know too much about the method,. Note that we are using a protected
        member of Worksheet ws._tables
        '''
        quoteRange = [(1, 1), (13, 5)]
        noteRange = [(1, 6), (13, 24)]
        quoteStyle = 'normStyle'
        noteStyle = 'explain'
        sumSize = 25
        margin = 25
        inputlen = 50   #len(df)

        wb = Workbook()
        ws = wb.active
        tf = TradeFormat(wb)

        # Make sure the out dir exists
        if not os.path.exists("out/"):
            os.mkdir("out/")

        # Make sure the file we are about to create does not exist
        dispath = "out/SCHNOrK.xlsx"
        if os.path.exists(dispath):
            os.remove(dispath)

        ls = LayoutSheet(sumSize, margin, inputlen)
        ls.styleTop(ws, 13, tf)

        wb.save(dispath)

        wb2 = load_workbook(dispath)
        ws2 = wb2.active

        x = ws2.merged_cells.ranges
        listOfMerged = list()
        listOfMerged.append(c((quoteRange[0])) + ':' +  c((quoteRange[1])))
        listOfMerged.append(c((noteRange[0])) + ':' +  c((noteRange[1])))
        for xx in x:
            # print (str(xx) in listOfMerged)
            assert str(xx) in listOfMerged
        assert ws[c(quoteRange[0])].style == quoteStyle
        assert ws[c(noteRange[0])].style == noteStyle

        assert len(ws._tables) == 1


        begin = c((1, ls.topMargin))

        end = c((13, ls.topMargin + ls.inputlen))
        tabRange = f'{begin}:{end}'
        assert tabRange == ws._tables[0].ref

        os.remove(dispath)

def notmain():
    '''Run some local code'''
        # pylint: disable = E1120
    ttt = TestLayoutSheet()
    # ttt.test_createImageLocation()
    # ttt.test_createWorkbook()
    ttt.test_styleTopwithnothin()




if __name__ == '__main__':
    notmain()
