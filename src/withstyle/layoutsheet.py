'''
Created on Oct 18, 2018

@author: Mike Petersen
'''
import os
import sqlite3

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, colors

# from inspiration.inspire import Inspire
from inspiration.inspire import Inspire
from journal.dfutil import DataFrameUtil
from journal.tradeutil import FinReqCol
from journal.xlimage import XLImage
from withstyle.tradestyle import c as tcell
from withstyle.tradestyle import style_range
from withstyle.thetradeobject import TheTradeObject, SumReqFields 



class LayoutSheet(object):
    '''
    Contains methods to layout the material on the excel page. Uses both pandas and openpyxl depending on
    how far in the program.  Generally the program progresses from pandas to openpyxl 
    '''


    def __init__(self, summarySize, topMargin, inputlen):
        '''
        Constructor
        :params:summarySize: The number of rows for each trade summary
        :params:frq: The FinReqCol object
        '''
        self.summarySize = summarySize
        self.topMargin = topMargin
        self.inputlen = inputlen
        
    def createImageLocation(self, df, ldf) :
        '''
        Debated what object to place this in. We are creating the shape of the excel document by adding rows
        to place the trade summaries and images, but we are doing it in a DataFrame and using  a list of 
        DataFrames, so it belongs here.
        Additionally creating the ImageLocation data structure to navigate the new space.
        '''
        #Add rows and append each trade, leaving space for an image. Create a list of names and row numbers 
        # to place images within the excel file (imageLocation data structure).
        
        frq = FinReqCol()
        newdf = DataFrameUtil.createDf(df,  self.topMargin)
    
        df = newdf.append(df, ignore_index = True)
        
        imageLocation = list()
        count=0
        for tdf in ldf :
            imageName='{0}_{1}_{2}_{3}.jpeg'.format (tdf[frq.tix].unique()[-1].replace(' ',''), 
               tdf[frq.name].unique()[-1].replace(' ','-'),
               tdf[frq.start].unique()[-1],
               tdf[frq.dur].unique()[-1])
            imageLocation.append([len(tdf) + len(df) + 3, 
                              tdf.Tindex.unique()[0].replace(' ', '') + '.jpeg',
                              imageName,
                              tdf.Start.unique()[-1],
                            tdf.Duration.unique()[-1]])
            print(count, imageName, len(imageLocation), len(tdf) + len(df) + 3)
            count = count + 1
            
            df = df.append(tdf, ignore_index = True)
            df = DataFrameUtil.addRows(df, self.summarySize)
        return imageLocation, df
 
    def createWorkbook(self, dframe) :
        nt = dframe
        # def 
        wb = Workbook()
        ws = wb.active
    
        for r in dataframe_to_rows(nt, index=False, header=False):
            ws.append(r)
    
        for name, cell  in zip(nt.columns, ws[self.topMargin]) :
            cell.value = name
        return wb, ws, nt
    
    
    def styleTop(self, ws, nt, tf) :
        ##Style the table, and the top paragraph.  Add and style the inspire quote. Create the SummaryMistake form (populate it below in a loop)
        tblRng= "{0}:{1}".format(tcell((1,self.topMargin)), tcell((len(nt.columns),self.topMargin + self.inputlen)))
        tab = Table(displayName="Table1", ref=tblRng)
        style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
    
        ws.add_table(tab)
        
        # A1:M5 inspire quote
        tf.mergeStuff(ws, (1,1), (13,5))
        ws["A1"].style = tf.styles["normStyle"]
        style_range(ws, "A1:M5", border=tf.styles["normStyle"].border)
        inspire = Inspire()
        ws["A1"] = inspire.getrandom().replace("\t", "        ")
        
        # A6:M24 introductory notes
        tf.mergeStuff(ws, (1,6), (13,24))
        ws["A6"].style = tf.styles["explain"]
        style_range(ws, "A6:M24", border=tf.styles["explain"].border)
        
        
        
        
        
        
        
        
    def createSummaries(self, imageLocation, ldf, jf, ws, tf) :
        tradeSummaries = list()
        XL = XLImage()
        srf = SumReqFields()
        
        response = input("Would you like to enter strategy names, targets and stops?     ")
        interview = True if response.lower().startswith('y') else False

        for loc, tdf in zip(imageLocation, ldf) :
            #     print('Copy an image into the clipboard for {0} beginning {1}, and lasting {2}'.format(loc[1], loc[2], loc[3]))
        
            
            img = XL.getAndResizeImage(loc[2], jf.outdir)
            
            #This is the location to place the chart on the page. Its kind of hidden in the deep recesses here.
            cellname = 'M' + str(loc[0])
            ws.add_image(img, cellname)
    
            #Put together the trade summary info for each trade and interview the trader
            tto=TheTradeObject(tdf, interview)
            tto.runSummary()
            tradeSummaries.append(tto.TheTrade)
    
            #Place the format shapes/styles in the worksheet
            tf.formatTrade(ws, anchor=(1, loc[0]))
    
            #populate the trade information
            for key in srf.tfcolumns.keys()  :
                cell = srf.tfcolumns[key][0]
                if isinstance(cell, list) :
                    cell = cell[0]
                tradeval = tto.TheTrade[key].unique()[0]
            #     print ("{0:10} \t{3} \t{1:}\t{2} ".format(key, cell, tradeval, tcell(cell, anchor=(1, loc[0]))))
    
    
                # Put some formulas in each trade Summary
                if key in srf.tfformulas :
    
                    anchor=(1,loc[0])
                    formula=srf.tfformulas[key][0]
                    args=[]
                    for c in srf.tfformulas[key][1:] :
                        args.append(tcell(c, anchor=anchor))
                    tradeval = formula.format(*args)
    
                if not tradeval :
                    continue
                ws[tcell(cell, anchor=(1, loc[0]))] = tradeval
                
        print("Done with interview")
        return tradeSummaries
            
    def createMistakeForm(self, tradeSummaries, mistake, ws, imageLocation) :
        '''
        Populate most of the mistake summaries form including the formulas retrieved from mistake. We
        do the cel translations here. Also the names are hyperlinked to the Trade Summary Form and the links
        are returned to the Mistake form, links are styled blue with double underline in both places.
        :params:tradeSummaries: A dataframe containing the the trade summaries info, one line per trade.
        :parmas:mistake: A dataframe containing the info to populate the mistake summary.
        :params:ws: The openpyxl worksheet object.
        :imageLocation: A list containing the locations in the worksheet for each of the trades in tradeSummaries.
        '''
        
        # Populate the name fields as hyperlinks
        for i in range(len(tradeSummaries)):
            key="name" + str(i+1)
            cell = mistake.mistakeFields[key][0][0]
            cell = tcell(cell, anchor=mistake.anchor)
            targetcell = (1, imageLocation[i][0])
            targetcell = tcell(targetcell)
            ts=tradeSummaries[i]
            cellval="{0} {1} {2}".format(i+1, ts.Name.unique()[0], ts.Account.unique()[0])
            link="#{}!{}".format(ws.title, targetcell)
            
            ws[cell].hyperlink=(link)
            ws[cell] = cellval
            ws[cell].font = Font(color=colors.WHITE, underline="double")
            
            link="#{}!{}".format(ws.title, cell)
            ws[targetcell].hyperlink = (link)
            ws[targetcell].font = Font(color=colors.WHITE, size=16, underline="double")
            
#             ft1 = Font(color=colors.RED,
#          size=11)
# ft2 = Font(color=colors.RED,
#          size=22,
#           italic=True)
# 
# a1.font=ft1
# d4.font=ft2
# 
# a1.font
# 
# a1.value="Its the end of the wolrd as we know it ..."
# d4.value="... And its about god damn time!"
# wb.save('out/fontStyle.xlsx')
            
            
            
    
        # Populate the pl (loss) fields and the mistake fields. These are all simple formulas.   
        tokens=["tpl", "pl", "mistake"]
        for token in tokens :
            for i in range(len(tradeSummaries)):
                key=token + str(i+1)
                if isinstance(mistake.mistakeFields[key][0], list) :
                    cell = mistake.mistakeFields[key][0][0]
                else :
                    cell = cell = mistake.mistakeFields[key][0]
                cell = tcell(cell, anchor=mistake.anchor)
            #     print(cell)
                formula = mistake.formulas[key][0]
                targetcell = mistake.formulas[key][1]
                targetcell = tcell(targetcell, anchor=(1, imageLocation[i][0]))
                formula = formula.format(targetcell)
    
                print("ws[{0}]='{1}'".format(cell,formula))
                ws[cell]=formula
                
    def createDailySummaryForm(self, TheTradeList, mistake, ws, anchor):
        '''
        Create the shape and populate the daily Summary Form
        :params:listOfTrade: A python list of the Summary Trade DataFrame, aka TheTrade, each one is a single row DataFrame
        containg all the data in the trade summaries.
        :params:mistke: 
        :params:ws: The openpyxl Worksheet object
        '''
        srf=SumReqFields()
        liveWins=list()
        liveLosses=list()
        simWins=list()
        simLosses=list()
        maxTrade = (0, "notrade")
        minTrade = (0, "notrade")
        #Didnot save the Trade number in TheTrade.  These should be the same order...
        count = 0
        
         
        for TheTrade in TheTradeList :
            pl = TheTrade[srf.pl].unique()[0]
            live = True if TheTrade[srf.acct].unique()[0] == "Live" else False
            count = count + 1
            if pl > maxTrade[0] :
                maxTrade = (pl, "Trade{0}, {1}, {2}".format(count, TheTrade[srf.acct].unique()[0], TheTrade[srf.name].unique()[0]))
            if pl < minTrade[0] :
                minTrade = (pl, "Trade{0}, {1}, {2}".format(count, TheTrade[srf.acct].unique()[0], TheTrade[srf.name].unique()[0]))
                
            if live:
                if pl > 0 :
                    liveWins.append(pl)
                else :
                    liveLosses.append(pl)
            else :
                if pl > 0 :
                    simWins.append(pl)
                else :
                    simLosses.append(pl)
                    
        anchor = (anchor[0], anchor[1] + mistake.numTrades + 5)    
        
        dailySumData = dict()   
        dailySumData['livetot'] = sum([sum(liveWins), sum(liveLosses)]) 
        
        numt = len(liveWins) + len(liveLosses)
        if numt == 0 : 
            dailySumData['livetotnote'] = "0 Trades"
        else :                                  
            dailySumData['livetotnote'] = "{0} Trade{1}, {2} Winner{3}, {4}, Loser{5}".format(numt, "" if numt  == 1 else "s", 
                                              len(liveWins), "" if len(liveWins) == 1 else "s",
                                              len(liveLosses), "" if len(liveLosses) == 1 else "s")      
        dailySumData['simtot']  = sum([sum(simWins), sum(simLosses)])
         
        # 9 trades,  3 Winners, 6 Losers
        numt=len(simWins) + len(simLosses)
        if numt == 0 : 
            dailySumData['simtotnote']  = "0 Trades"
        else :                                  #4 trades, 1 Winner, 3 Losers
            dailySumData['simtotnote'] = "{0} Trade{1}, {2} Winner{3}, {4}, Loser{5}".format(numt, "" if numt == 1 else "s", 
                                              len(simWins), "" if len(simWins) == 1 else "s",
                                              len(simLosses), "" if len(simLosses) == 1 else "s")      
        
               
        dailySumData['highest']     = maxTrade[0]  
        dailySumData['highestnote'] = maxTrade[1]
        dailySumData['lowest']      = minTrade[0]
        dailySumData['lowestnote']  = minTrade[1]
        dailySumData['avgwin']      = sum([sum(liveWins), sum(simWins)]) / (len(liveWins) + len(simWins))
        dailySumData['avgwinnote']  = "X {} =  ${:.2f}".format(len(liveWins) + len(simWins), sum([sum(liveWins), sum(simWins)]))
        dailySumData['avgloss']     = sum([sum(liveLosses), sum(simLosses)]) / (len(liveLosses) + len(simLosses))
        dailySumData['avglossnote'] = "X {} =  (${:.2f})".format(len(liveLosses) + len(simLosses), abs(sum([sum(liveLosses), sum(simLosses)])))
                                  
        for key in dailySumData.keys() :
            rng = mistake.dailySummaryFields[key][0]
            if isinstance(rng, list) :
                rng = rng[0]
            ws[tcell(rng, anchor=anchor)] = dailySumData[key] 

    
    
    
    def save(self, wb, jf):
        #Write the file
        jf.mkOutdir() 
        saveName=jf.outpathfile
        count=1
        while True :
            try :
                wb.save(saveName)
            except PermissionError as ex :
                print(ex)
                print("Failed to create file {0}.{1}".format(saveName, ex))
                print("Images from the clipboard were saved  in {0}".format(jf.outdir))
                (nm, ext) = os.path.splitext(jf.outpathfile)
                saveName = "{0}({1}){2}".format(nm,count,ext)
                print("Will try to save as {0}".format(saveName))
                count=count+1
                if count==6:
                    print("Giving up. PermissionError")
                    raise (PermissionError("Failed to create file {0}".format(saveName)))
                continue
            except Exception as ex:
                print (ex)
            break
        print("Done!")

                