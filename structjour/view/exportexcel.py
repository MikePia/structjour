# Structjour -- a daily trade review helper
# Copyright (C) 2019 Zero Substance Trading
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.
'''
Export to an excel document.

Created on May 29, 2019

@author: Mike Petersen
'''
import datetime as dt
import logging
import os
from PIL import Image as PILImage

from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors

import numpy as np
import pandas as pd

from PyQt5.QtCore import QSettings

from structjour.dfutil import DataFrameUtil
from structjour.colz.finreqcol import FinReqCol
from structjour.definetrades import DefineTrades
from structjour.thetradeobject import SumReqFields
from structjour.tradestyle import c as tcell
from structjour.layoutsheet import LayoutSheet
from structjour.dailysumforms import MistakeSummary
from structjour.view.dailycontrol import DailyControl
from structjour.xlimage import XLImage


from structjour.tradestyle import TradeFormat
# pylint: disable = C0103

class ExportToExcel:
    '''
    Export to excel from python objects
    '''

    def __init__(self, ts, jf, df):
        self.ts = ts
        self.jf = jf
        self.df = df
        self.topMargin = 25             # For use in Excel export

    def saveXL(self, wb, jf):
        '''
        Save wb as an excel file. If Permission Denied error is thrown, try renaming it.
        '''
        #Write the file
        jf.mkOutdir()
        saveName = jf.outpathfile
        count = 1
        saveName = saveName.replace('/', '\\')
        while True:
            try:
                wb.save(saveName)
            except PermissionError as ex:
                logging.error(ex)
                logging.error("Failed to create file {0}".format(saveName))
                logging.info("Images from the clipboard were saved  in {0}".format(jf.outdir))
                (nm, ext) = os.path.splitext(jf.outpathfile)
                saveName = "{0}({1}){2}".format(nm, count, ext)
                logging.error("Will try to save as {0}".format(saveName))
                count = count+1
                if count == 6:
                    raise (PermissionError(
                        "Failed to create file {0}".format(saveName)))
                continue
            except Exception as ex:
                logging.error(ex)
            break

    def populateXLDailyNote(self, ws):
        '''
        Export the daily note as found in the DailyControl dialog.
        '''
        cell = (1, 6)
        cell = tcell(cell)
        dc = DailyControl(self.jf.theDate)
        note = dc.getNote()
        if note:
            ws[cell] = note

    def populateXLDailySummaryForm(self, mistake, ws, anchor):
        '''
        For the export to excel daily summary forms.
        Populate the daily Summary Form. The values are retrieved from self.ts via the
        DailyControl dialog object-without calling runDialog. The static labels are set earlier
        at the creation of the form shape/style. This method sets some statistics and notes for
        things like  average winners/losers etc.

        :params mistke:
        :params ws: The openpyxl Worksheet object
        :raise Value Error: When pl is misformatted and cannot be used.
        '''

        dc = DailyControl()
        dailySumData = dc.gatherDSumData(self.ts)
        anchor = (anchor[0], anchor[1] + mistake.numTrades + 5)
        for key in dailySumData.keys():
            cell = mistake.dailySummaryFields[key][0]
            if isinstance(cell, list):
                cell = cell[0]
            ws[tcell(cell, anchor=anchor)] = dailySumData[key]

    def populateXLMistakeForm(self, mistake, ws, imageLocation):
        '''
        For the export to excel the mistake summary form. Populate the dynamic parts of mistake
        summaries. That includes fomulas and hyperlinks with references to tradeSummaries and
        hyperlinks to back. The cell location for the links to tradeSummaries requires then
        anchor information in imageLocation and the specific cell within tradeSummaries found in
        mistakeFields. The return hyperlinks in the tradeSummaries forms are also translated here.

        :parmas mistake: A dataframe containing the info to populate the mistake summary form.
        :params ws: The openpyxl worksheet object.
        :parmas imageLocation: A list containing the locations in the worksheet for each of the
                               trades in tradeSummaries.
        '''

        # Populate the name fields as hyperlinks to tradeSummaries title cell and back.
        for i, (iloc, tradekey) in enumerate(zip(imageLocation, self.ts)):
            tsum = self.ts[tradekey]
            key = "name" + str(i+1)
            cell = mistake.mistakeFields[key][0][0]
            cell = tcell(cell, anchor=mistake.anchor)
            targetcell = (1, iloc[0][0][1])
            targetcell = tcell(targetcell)
            cellval = "{0} {1} {2}".format(i+1, tsum.Name.unique()[0], tsum.Account.unique()[0])
            link = "#{}!{}".format(ws.title, targetcell)

            ws[cell].hyperlink = (link)
            ws[cell] = cellval
            ws[cell].font = Font(color=colors.WHITE, underline="double")

            link = "#{}!{}".format(ws.title, cell)
            ws[targetcell].hyperlink = (link)
            ws[targetcell].font = Font(color=colors.WHITE, size=16, underline="double")

        # Populate the pl (loss) fields and the mistake fields. These are all formulas like =B31
        tokens = ["tpl", "pl", "mistake"]
        for token in tokens:
            for i in range(len(self.ts)):
                key = token + str(i+1)
                if isinstance(mistake.mistakeFields[key][0], list):
                    cell = mistake.mistakeFields[key][0][0]
                else:
                    cell = cell = mistake.mistakeFields[key][0]
                cell = tcell(cell, anchor=mistake.anchor)
                formula = mistake.formulas[key][0]
                targetcell = mistake.formulas[key][1]
                targetcell = tcell(targetcell, anchor=(1, imageLocation[i][0][0][1]))
                formula = formula.format(targetcell)

                # print("ws[{0}]='{1}'".format(cell, formula))
                ws[cell] = formula

    def placeImagesAndSumStyles(self, imageLocation, ws, tf):
        '''
        A helper method for export to excel. Place images and the tradesummary stylesin the
        openpyxl ws object
        '''

        # tradeSummaries = list()
        CELLS = 20 #trial and error here
        srf = SumReqFields()

        for loc in imageLocation:
            #Place the format shapes/styles in the worksheet
            tf.formatTrade(ws, srf, anchor=(1, loc[0][0][1]))
            for iloc, fn in zip(loc[0], loc[1]):
                if not os.path.exists(fn):
                    continue
                img = PILImage.open(fn)
                xl = XLImage()
                newSize = xl.adjustSizeByHeight(img.size, CELLS)
                img = img.resize(newSize, PILImage.ANTIALIAS)
                img.save(fn, 'png')
                img = Image(fn)
                if img:
                    cellname = tcell(iloc)
                    ws.add_image(img, cellname)


    def populateXLDailyFormulas(self, imageLocation, ws):
        '''
        Helper method for export to excel. Populate the excel formulas in the daily summary forms.
        The formulas are in self.ts and the information to transalte cell location is in
        imageLocation
        '''
        srf = SumReqFields()
        for loc, tradekey in zip(imageLocation, self.ts):
            tto = self.ts[tradekey]

            #populate the trade information
            for key in srf.tfcolumns:
                cell = srf.tfcolumns[key][0]
                if isinstance(cell, list):
                    cell = cell[0]
                tradeval = tto[key].unique()[0]

                # Put some formulas in each trade Summary
                if key in srf.tfformulas:

                    anchor = (1, loc[0][0][1])
                    formula = srf.tfformulas[key][0]
                    args = []
                    for c in srf.tfformulas[key][1:]:
                        args.append(tcell(c, anchor=anchor))
                    tradeval = formula.format(*args)

                if not tradeval:
                    continue
                if isinstance(tradeval, (pd.Timestamp, dt.datetime, np.datetime64)):
                    tradeval = pd.Timestamp(tradeval)
                elif isinstance(tradeval, bytes):
                    tradeval = None


                ws[tcell(cell, anchor=(1, loc[0][0][1]))] = tradeval

        return

    def getImageNamesFromTS(self):
        '''
        Gathers specific image names as stored in self.ts[tradekey][chartkey]
        :return: A dict of all image names in self.ts organized by trade with the keys 'Trade n'
            These keys are shared by the index names of Tindex column of the statement DataFrame
        '''
        imageNames = dict()
        for key in self.ts:
            images = []
            keys = self.ts[key].keys()
            charts = ['chart1', 'chart2', 'chart3']
            for chart in charts:
                if chart in keys:
                    if os.path.exists(self.ts[key][chart].unique()[0]):
                        images.append(self.ts[key][chart].unique()[0])
            key = key.split(' ')
            newkey = 'Trade ' + key[0]
            imageNames[newkey] = images
        return imageNames

    def layoutExcelData(self, df, ldf, imageNames):
        '''
        1) Determine the locations in the Excel doc to place trade summaries, trade tables and
            images.
        2) Create the empty rows and place the trade tables in the df according to the locations.
        :params df: We requre the df as  a whole because we are adding rows to it.
        :params ldf: A list of dataframes, each a trade, each one is placed into our new skeletal
                     layout for excel
        :return (Imagelocation, df): ImageLocation contains
                                [ [list of image location],   # up to 3 per trade
                                  [list of image names],      # up to 3 per trade
                                  Start time,
                                  trade dur,
                                ]
        '''

        imageLocation = list()
        srf = SumReqFields()
        sumSize = srf.maxrow() + 5
        summarySize = sumSize
        spacing = 3

        # Image column location
        c1col = 13
        c2col = 1
        c3col = 9
        frq = FinReqCol()
        newdf = DataFrameUtil.createDf(df, self.topMargin)

        df = newdf.append(df, ignore_index=True)

        for tdf in ldf:
            theKey = tdf[frq.tix].unique()[-1]
            imageName = imageNames[theKey]
            xtraimage = 0                       # Add space for second/third image
            if len(imageName) > 1:
                xtraimage = 21
            ilocs = []
            #Need 1 entry even if there are no images
            ilocs.append((c1col, len(tdf) + len(df) + spacing))
            for i in range(0, len(imageName)):
                if i == 1:
                    ilocs.append((c2col, len(tdf) + len(df) + spacing + 20))
                elif i == 2:
                    ilocs.append((c3col, len(tdf) + len(df) + spacing + 20))

            # Holds image locations, image name, trade start time, trade duration as delta
            imageLocation.append([ilocs,
                                  imageName,
                                  tdf.Start.unique()[-1],
                                  tdf.Duration.unique()[-1]])

            # Append the mini trade table then add rows to fit the tradeSummary form
            df = df.append(tdf, ignore_index=True)
            df = DataFrameUtil.addRows(df, summarySize + xtraimage)
        return imageLocation, df

    def exportExcel(self):
        '''
        Export to excel the trade tables, trade summaries, and daily forms
        '''

        # Create the space in dframe to add the summary information for each trade.
        # Then create the Workbook.
        settings = QSettings('zero_substance', 'structjour')
        val = settings.value('inputType')

        # Get a list of Trades from self.df
        tu = DefineTrades(val)
        ldf = tu.getTradeList(self.df)

        # Lay out a dataframe with space for charts
        imageNames = self.getImageNamesFromTS()
        imageLocation, dframe = self.layoutExcelData(self.df, ldf, imageNames)
        assert len(ldf) == len(imageLocation)

        # Create an openpyxl wb from the dataframe
        ls = LayoutSheet(self.topMargin, len(self.df))
        wb, ws, nt = ls.createWorkbook(dframe)

        tf = TradeFormat(wb)
        ls.styleTop(ws, len(nt.columns), tf)

        mstkAnchor = (len(dframe.columns) + 2, 1)
        mistake = MistakeSummary(numTrades=len(ldf), anchor=mstkAnchor)
        mistake.mstkSumStyle(ws, tf, mstkAnchor)
        mistake.dailySumStyle(ws, tf, mstkAnchor)

        # tradeSummaries = ls.runSummaries(imageLocation, ldf, self.jf, ws, tf)
        self.placeImagesAndSumStyles(imageLocation, ws, tf)
        self.populateXLDailyFormulas(imageLocation, ws)

        self.populateXLMistakeForm(mistake, ws, imageLocation)
        self.populateXLDailySummaryForm(mistake, ws, mstkAnchor)
        self.populateXLDailyNote(ws)


        self.saveXL(wb, self.jf)
        logging.info("Processing complete. Saved {}".format(self.jf.outpathfile))
        # return jf
