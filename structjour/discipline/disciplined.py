# Structjour -- a daily trade review helper
# Copyright (C) 2019 Zero Substance Trading
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301, USA.

import datetime as dt
import os
from openpyxl import load_workbook


from structjour.definetrades import DefineTrades
from structjour.dfutil import DataFrameUtil
from structjour.statements.findfiles import getDirectory
from structjour.statements.ibstatementdb import StatementDB
from structjour.thetradeobject import SumReqFields
from structjour.tradestyle import c as tcell

import pandas as pd
# pylint: disable=C0103


class DisReqCol(object):
    def __init__(self, thedate=dt.date.today()):
        self.thedate = thedate
        rcvals = ["Date", "Time", "Long/Short", "Ticker", "Entry Price", "Account Balance", "Position",
                  "Stop Loss", "Target", "Average Exit", "PnL", "Strategy", "Trade Notes", 'TS ID']

        rckeys = ["date", "time", "side", "symb", "entry1", "acctbal", "shares",
                  "stoploss", "targ", "avgexit", "pl", "strat", "notes", 'id']

        rc = dict(zip(rckeys, rcvals))

        self._date = rc['date']
        self._time = rc['time']
        self.side = rc['side']
        self.symb = rc['symb']
        self.entry1 = rc['entry1']
        self.acctbal = rc['acctbal']
        self.shares = rc['shares']
        self.stoploss = rc['stoploss']
        self.target = rc['targ']
        self.avgexit = rc['avgexit']
        self.pl = rc['pl']
        self.strat = rc['strat']
        self.notes = rc['notes']
        self.id = rc['id']

        # We really don't need to manage the styles but just let these
        # sit here for uniformity-- or finding we actually do need them.
        tfcolumns = {
            "date": [(1, 1), 'dateStyle'],
            "time": [(2, 1), 'timeStyle'],
            "side": [(3, 1), 'normal'],
            "symb": [(4, 1), 'normal'],
            "entry1": [(5, 1), 'currencyStyle'],
            "acctbal": [(6, 1), 'currencyStyle'],
            "shares": [(7, 1), 'numberStyle'],
            "stoploss": [(8, 1), 'numberStyle'],
            "targ": [(9, 1), 'currencyStyle'],
            "avgexit": [(10, 1), 'currencyStyle'],
            "pl": [(11, 1), 'currencyStyle'],
            "strat": [(12, 1), 'normal'],
            "notes": [(16, 1), 'normal'],
            "id": [(17, 1), 'normal']}

        self.rc = rc
        self.tfcolumns = tfcolumns

def doctorTheTrade(newdf, daDate, ts,  fname):
    '''
    Fix up the time entries to contain the dates as Timestamps and charts entries. Plan not to try
    to deal with the whole excel related files. The specs for which images go where are non-
    existent as near as I can tell and I really hate dealing with MS files directly. Instead, load
    up images as saved in the directory by trade index names if found for 1 chart and place strings
    in chart2 and chart3
    '''
    for i in range(0, 8):
        key = 'Time' + str(i+1)
        if newdf[key].any():
            val = newdf[key].unique()[0]
            datestring = daDate.strftime('%Y-%m-%d ')
            if isinstance(val, str):
                if len(val) == 8 or len(val.split(':')) == 3:
                    d = pd.Timestamp(datestring + val)
                    newdf[key] = d
                else:
                    raise ValueError('Unrecognized timestring:', val)
            elif isinstance(val, (dt.datetime, dt.time)):
                timestring = val.strftime('%H:%M:%S')
                newdf[key] = pd.Timestamp(datestring + timestring)
            elif isinstance(val, pd.Timestamp):
                pass
            else:
                msg = 'Unexpected type: ' + str(type(val))
                raise ValueError(msg)
        else:
            pass
            # print(newdf[key].unique())
            # print(len(newdf[key].unique()))
    outdir, n = os.path.split(fname)
    for key in ts:
        tto = ts[key]
        thechart = None
        if not os.path.exists(outdir):
            break
        filelist = os.listdir(outdir)
        for fn in filelist:
            num = key.split(' ')[0]
            tradekey = f'trade{num}_'
            if tradekey in fn.lower():
                '''We have to bring the fromat into the the 2019s and add the keys we now use'''
                thechart = os.path.join(outdir, fn)
                now = pd.Timestamp.today()
                ts[key]['chart1'] = fn
                ts[key]['chart1Start'] = now
                ts[key]['chart2Start'] = now
                ts[key]['chart3Start'] = now
                ts[key]['chart1End'] = now
                ts[key]['chart2End'] = now
                ts[key]['chart3End'] = now
                ts[key]['chart1Interval'] = 1
                ts[key]['chart2Interval'] = 5
                ts[key]['chart3Interval'] = 15
                ts[key]['chart2'] = 'The animated parrot chart'
                ts[key]['chart3'] = 'The animated parrot complaint deptment chart'
                break
    return newdf

def loadTradeSummaries(loc, trades):
    '''
    Load up each trade summary in the excel doc into a 1 row DataFrame, return a list
    of these DataFrames. The addresses are supplied by srf plus the loc object that has the
    anchors.
    :params:loc: A list of the rows within the excel doc on which to find the trade summaries
    :return: A list of 1-row DataFrames Each trade is on one row from each of the trade summay forms
    '''

    ldf = list()
    ts = dict()
    srf = SumReqFields()
    reqCol = srf.rc
    newdf = pd.DataFrame(columns=reqCol.values())
    colFormat = srf.tfcolumns

    for i, rowNum in enumerate(loc):
        newdf = DataFrameUtil.createDf(newdf, 1)
        for key in reqCol.keys():
            if key in ['date', 'clean', 'id']:
                continue
            cell = colFormat[reqCol[key]][0]
            if isinstance(cell, list):
                cell = cell[0]
            cell = tcell(cell, anchor=(1, rowNum))
            newdf.iloc[-1][reqCol[key]] = trades[cell].value

        tradekey = str(i+1) + ' ' +  newdf.iloc[0].Name
        ts[tradekey] = newdf
        ldf.append(newdf)
    return ldf, ts

def getTradeSummaryFormLocations(ws):
    '''
    Find the Trade Summary Forms within the excel doc and save their location in tradeLoc 
    trades.max_row
    :params:ws: The openpyxl Worksheet object to search
    '''
    loc = list()

    #Search for trade names like 'AAPL Short' in column A. Save the cell for each to loc
    for row in ws.iter_rows():
        val = row[0].value
        if val and len(val) > 3 and len(val) < 15:
            if 'short' in val.lower() or 'long' in val.lower():
                r = row[0].row
                loc.append(r)
    return loc

def getAvgExit(trade):
    '''We want to display jut one exit in the Trade log. It might as well be the average
    exit when we scale out. If there are more than 8 combined entries and exits--could be
    a problem
    TODO: track down the note for more entrances/exits than 8 -- It may be complicated-- 
    entered a bug--classic procrastination
    '''
    product = 0
    exits = list()
    positions = list()
    for i in range(1, 9):
        ex = "Exit{}".format(i)
        shares = "EShare{}".format(i)
        if trade[ex].unique()[0]:
            x = trade[ex].unique()[0]
            p = trade[shares].unique()[0]
            exits.append(x)
            positions.append(p)
            try:
                assert(positions[-1])
            except:
                print(
                    "Found empty shares in an exit-- this may be an overnight trade or something: Fix this bug developer!!")
            product = product + (x * p)

    return 0 if sum(positions) == 0 else product/sum(positions)


def gotAnyExits(trade):
    '''
    Return True if this trade processed at least one exit. Return False if the exits are all None
    :params:trade: DataFrame representing a single trade and with columns Exit1, Exit2 ... Exit8
    '''
    for i in range(1, 9):
        ex = "Exit{}".format(i)
        if trade[ex].unique()[0]:
            return True
    return False

def getWeekCount(theDate):
    '''
    Returns the number of weeks that have happend in the month counting only weekdays. If the first day
    of the month is Friday, the following Monday begins week 2. If the first day of the week is Saturday
    Week 2 will begin on Monday the 10th.
    :params:theDate: the date to test. A timedate.date object
    '''
    startMonth = dt.date(theDate.year, theDate.month, 1)
    delt = dt.timedelta(1)
    weekCount = 1
    startCounting = False
    while startMonth <= theDate:

        if(startMonth.weekday() == 4):
            startCounting = True
        if startCounting == True and startMonth.weekday() == 0:
            weekCount = weekCount + 1
        startMonth = startMonth+delt
    return weekCount

def getDevelDailyJournalList(prefix, begin):
    '''
    
    :return: A list of filenames with dates[[filename, date], [filename2, date2]]...
    '''
    thelist = list()
    current = pd.Timestamp(begin)
    now = pd.Timestamp.today()
    delt = pd.Timedelta(days=1)
    while current < now:
        outdir = getDirectory(current)
        outdir = os.path.join(outdir, 'out')
        fn = current.strftime("trades%A_%m%d.xlsx")
        fn = os.path.join(outdir, fn)
        if os.path.exists(fn):
            thelist.append([fn, current])
            # print(current.strftime('%A, %B %d, %Y'), fn)
        current += delt
    return thelist

def recreateFPEntries(ts):
    fpentries = dict()
    entry = list()
    for key in ts:
        entry = list()
        for i in range(1,9):
            price = side = dtime = None
            entrykey = 'Entry' + str(i)
            exitkey = 'Exit' + str(i)
            sharekey = 'EShare' + str(i)
            timekey = 'Time' + str(i)

            if ts[key][entrykey].any():
                price = ts[key][entrykey].item()
            elif ts[key][exitkey].item():
                price = ts[key][exitkey].item()

            if ts[key][sharekey].item():
                side = 'B' if ts[key][sharekey].item() > 0 else 'S'

            if ts[key][timekey].item():
                dtime = ts[key][timekey].item()
            if price == None and side == None and dtime == None:
                break
            entry.append([price, 'deprecated', side, dtime])
        fpentries[key] = entry
    return fpentries

def getTradeSummary(fname, daDate):
    '''
    Top level script to load the trades in fname.
    :fname: An excel file generated by structjour containing a Daily trade table, a Daily notes
    field, A mistake summary form, a daily summary form and a trade summary form for each trade.
    :return: (ldf, ts, fpentries) ldf is a list of DataFrames, each containg the transactions for
            a trade. ts is a Dictionary containng the Trade summary information, including user
            input for things like target and stop loss and names of charts. fpentries, a dict 
            with the entries for each trade. ts and fpentries share the same keys.
    ''' 
    wb = load_workbook(fname)
    ws = wb.active
    tradeLoc = getTradeSummaryFormLocations(ws)
    ldf, ts = loadTradeSummaries(tradeLoc, ws)
    fpentries = recreateFPEntries(ts)
    for tdf in ldf:
        tdf = doctorTheTrade(tdf, daDate, ts, fname) 
    return ldf, ts, fpentries

def doctorThetable(df, daDate):
    '''Add a Date field if its missing'''

    if 'Date' not in df.columns:
        d = daDate
        for i, row in df.iterrows():
            if not row.Time:
                break
            t = row.Time
            assert isinstance(t, str)
            datestring = daDate.strftime('%Y-%m-%d ')
            if len(t) == 8 and len(t.split(':')) == 3:
                newd = pd.Timestamp(datestring + t)
                df.at[i, 'Date'] = newd
            else:
                raise ValueError('Unrecognized timestring:', t)

    return df

def getTradeTable(fname, daDate):
    wb = load_workbook(fname)
    ws = wb.active
    
    data = list()
    columns = list()
    start = False
    tot1 = False
    collen = 0
    notes = None
    gotsomething = False
    gotitforsure = False
    for row in ws.iter_rows():
        val = row[0].value

        # This should get the first item above the table if there is only one and the second item
        # if there are two above the table. IOW, this should get the quote or the notes-if they
        # exist- with a preference for the notes.
        if not start and not tot1:
            if not gotsomething and val != 'Tindex':
                if val:
                    notes = val
                    gotsomething = True
            elif val and gotitforsure == False and val != 'Tindex':
                notes = val
                gotitforsure = True

        darow = list()
        if val and val == 'Tindex':
            start = True
            for cval in row:
                if cval.value == None:
                    collen = len(columns)
                    break
                columns.append(cval.value)
        elif start:
            if row[0].value and row[0].value.startswith('Trade'):
                for i in range(0, collen):
                    darow.append(row[i].value)
                data.append(darow)
            else:
                for i in range(0, collen):
                    darow.append(row[i].value)
                data.append(darow)
                tot1 = True
                start = False
        elif tot1:
            for i in range(0, collen):
                    darow.append(row[i].value)
            data.append(darow)
            break
    dframe = pd.DataFrame(data=data, columns=columns)
    dframe = doctorThetable(dframe, daDate)
    return dframe, notes

def registerTradesOLD(tsList, wb):
    for fname, theDate in tsList:
        # print(fname, theDate)
        wb2 = load_workbook(fname)
        trades = wb2["Sheet"]

        tradeLocation = getTradeSummaryFormLocations(trades)

        ldf, ts = loadTradeSummaries(tradeLocation, trades)
        
        drc = DisReqCol(theDate)

        tlog = wb["Trade Log"]

        startSearch = False
        ix = -2
        cols = drc.tfcolumns
        # Here is a list of the keys to use cols.keys() of the trade log DataFrame
        #['date', 'time', 'side', 'symb', 'entry1', 'acctbal', 'shares',
        #'stoploss', 'targ', 'avgexit', 'pl', 'notes'])
        # Not bothering with the abstraction (drc.name) because this is entirely ours.

        srf = SumReqFields()

        for row in tlog.iter_rows():
            anchor = (1, row[0].row)

            if startSearch == True:
                if not row[0].value:
                    startSearch = False
                    ix = 0
            if row[0].value == 'Date':
                startSearch = True

            if ix >= 0:
                tdf = ldf[ix]
                if not gotAnyExits(tdf):
                    continue
                # print(cols.keys())

                #date
                cell = tcell(cols['date'][0], anchor=anchor)
                tlog[cell] = theDate

                #time
                cell = tcell(cols['time'][0], anchor=anchor)
                tim = tdf[srf.start].unique()[0]
                tlog[cell] = tim

                #side
                name = tdf[srf.name].unique()[0]
                if name:
                    cell = tcell(cols['side'][0], anchor=anchor)
                    tlog[cell] = name.split()[1]

                #symb
                cell = tcell(cols['symb'][0], anchor=anchor)
                tlog[cell] = name.split()[0]

                #entry1
                cell = tcell(cols['entry1'][0], anchor=anchor)
                tlog[cell] = tdf[srf.entry1].unique()[0]

                #Account Balance (setting an excel formula)
                cell = tcell(cols['acctbal'][0], anchor=anchor)
                formula = "=$M$3+SUM($N$7:$N{})".format(row[0].row-1)
                tlog[cell] = formula

                # "shares"
                cell = tcell(cols['shares'][0], anchor=anchor)
                shares = tdf[srf.shares].unique()[0].split()[0]
                if len(shares) > 0:
                    try:
                        ishares = int(shares)
                    except:
                        ishares = 0
                tlog[cell] = ishares

                #stoploss
                cell = tcell(cols['stoploss'][0], anchor=anchor)
                sl = tdf[srf.stoploss].unique()[0]
                tlog[cell] = sl

                #target
                cell = tcell(cols['targ'][0], anchor=anchor)
                target = tdf[srf.targ].unique()[0]
                tlog[cell] = target

                #avgExit
                cell = tcell(cols['avgexit'][0], anchor=anchor)
                tlog[cell] = getAvgExit(tdf)

                # P/L
                cell = tcell(cols['pl'][0], anchor=anchor)
                pl = tdf[srf.pl].unique()[0]
                tlog[cell] = pl

                # Strategy
                cell = tcell(cols['strat'][0], anchor=anchor)
                strat = tdf[srf.strat].unique()[0]
                tlog[cell] = strat


                # notes (from the mistake note field)
                cell = tcell(cols['notes'][0], anchor=anchor)
                note = tdf[srf.mstknote].unique()[0]
                tlog[cell] = note

                ix = ix + 1
                if ix == len(ldf):
                    break

def registerTrades(wb, theDate):

    ibdb = StatementDB()
    df = ibdb.getStatement(theDate)
    x = ibdb.getNumTicketsforDay(theDate)
    if not x[0] or not x[1]:
        msg = f'''found {x[0]} tickets and {x[1]} trades. Nothing to process for the date {theDate.strftime('%A %d, %Y')}in the DB'''
        return msg
    tu = DefineTrades('DB')
    inputlen, df, ldf = tu.processDBTrades(df)

    ts, entries = ibdb.getTradeSummaries(theDate)
    # self.ts = setTradeSummaryHeaders(ts)
    
        
    drc = DisReqCol(theDate)

    tlog = wb["Trade Log"]
    account = tlog['I1'].value

    startSearch = False
    ix = -2
    cols = drc.tfcolumns
    # Here is a list of the keys to use cols.keys() of the trade log DataFrame
    #['date', 'time', 'side', 'symb', 'entry1', 'acctbal', 'shares',
    #'stoploss', 'targ', 'avgexit', 'pl', 'notes'])
    # Not bothering with the abstraction (drc.name) because this is entirely ours.

    srf = SumReqFields()

    ids = list()

    for row in tlog.iter_rows():
        #Unconventional loop stuff 
        anchor = (1, row[0].row)
        keys = list(ts.keys())
        if startSearch == True:
            idcol = cols['id'][0][0] - 1
            if row[idcol].value:
                ids.append(row[idcol].value)
            if not row[0].value:
                startSearch = False
                ix = 0

        if row[0].value == 'Date':
            startSearch = True
        

        if ix >= 0 and ix < len(keys):
            
            while True and ix < len(keys):
                key = keys[ix]
                tdf = ts[key]
                if not gotAnyExits(tdf) or tdf['id'].unique()[0] in ids or (
                    tdf['Account'].unique()[0] != account):
                    if tdf['Account'].unique()[0] not in ['SIM', 'Live']:
                        raise ValueError('Programmer exception in search of weird data')
                    # Continue the inner loop -- keep row from the outer loop loop
                    ix += 1
                    continue

                #date
                cell = tcell(cols['date'][0], anchor=anchor)
                tlog[cell] = theDate

                #time
                cell = tcell(cols['time'][0], anchor=anchor)
                tim = tdf[srf.start].unique()[0]
                tlog[cell] = tim

                #side
                name = tdf[srf.name].unique()[0]
                if name:
                    cell = tcell(cols['side'][0], anchor=anchor)
                    tlog[cell] = name.split()[1]

                #symb
                cell = tcell(cols['symb'][0], anchor=anchor)
                tlog[cell] = name.split()[0]

                #entry1
                cell = tcell(cols['entry1'][0], anchor=anchor)
                tlog[cell] = tdf[srf.entry1].unique()[0]

                #Account Balance (setting an excel formula)
                cell = tcell(cols['acctbal'][0], anchor=anchor)
                formula = "=$M$3+SUM($N$7:$N{})".format(row[0].row-1)
                tlog[cell] = formula

                # "shares"
                cell = tcell(cols['shares'][0], anchor=anchor)
                shares = tdf[srf.shares].unique()[0].split()[0]
                if len(shares) > 0:
                    try:
                        ishares = int(shares)
                    except:
                        ishares = 0
                tlog[cell] = ishares

                #stoploss
                cell = tcell(cols['stoploss'][0], anchor=anchor)
                sl = tdf[srf.stoploss].unique()[0]
                if isinstance(sl, bytes):
                    sl = None
                tlog[cell] = sl

                #target
                cell = tcell(cols['targ'][0], anchor=anchor)
                target = tdf[srf.targ].unique()[0]
                if isinstance(target, bytes):
                    target = None
                tlog[cell] = target

                #avgExit
                cell = tcell(cols['avgexit'][0], anchor=anchor)
                tlog[cell] = getAvgExit(tdf)

                # P/L
                cell = tcell(cols['pl'][0], anchor=anchor)
                pl = tdf[srf.pl].unique()[0]

                # Don't know how these are sneaking through-- Is it only on legacy files?
                if isinstance(pl, bytes):
                    pl = 0.0
                tlog[cell] = pl

                # Strategy
                cell = tcell(cols['strat'][0], anchor=anchor)
                strat = tdf[srf.strat].unique()[0]
                tlog[cell] = strat


                # notes (from the mistake note field)
                cell = tcell(cols['notes'][0], anchor=anchor)
                note = tdf[srf.mstknote].unique()[0]
                tlog[cell] = note

                # id (from the database)
                cell = tcell(cols['id'][0], anchor=anchor)
                id = tdf['id'].unique()[0]
                tlog[cell] = id

                # break the inner loop
                break


            ix += 1
            if ix >= len(keys):
                # done with outer loop
                break
    return None



def local():
    daDate = pd.Timestamp('20181217')
    fn = 'c:/trader/journal/_201812_December/_1217_Monday/out/tradesMonday_1217.xlsx'
    if not os.path.exists(fn):
        print('try again')
    else:
        ldf, ts, fpentries = getTradeSummary(fn, daDate)

def notmain():

    import sys
    disPath = os.path.normpath("C:/trader/journal/Disciplined.xlsx")
    if os.path.exists(disPath):
        wb = load_workbook(disPath)
    else:
        print("Does not exist", disPath)
        sys.exit(0)


    begin = pd.Timestamp(2018, 11, 5)
    prefix = "C:/trader/journal/"
    # flist = getDevelDailyJournalList(prefix, begin)
    current = begin
    now = pd.Timestamp.today()
    delt = pd.Timedelta(days=1)
    while now > current:
        registerTrades(wb, current)
        print('Processed date', current.strftime('%A, %B %d, %Y'))
        current += delt
    wb.save(disPath)
    print('done!')




if __name__ == '__main__':
    # local()
    notmain()